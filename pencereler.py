# -*- coding: utf-8 -*-
"""Yardimci pencereler: gecmis taramalari listeleyen GecmisPenceresi ve
disaridan gelen bir listeyle karsilastirma sonucunu gosteren
KarsilastirmaPenceresi.

desktop.py'den ayrildi -- ikisi de QDialog alt siniflari, MainWindow'a
sadece `ebeveyn` referansiyla (opsiyonel geri-cagrilar icin) bagli."""

import os
import tempfile
from datetime import datetime
from collections import Counter

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QMessageBox, QButtonGroup,
)

import gecmis
import karsilastirma
from gorsel_araclari import dosyayi_sistemde_ac


class KarsilastirmaPenceresi(QDialog):
    """Dış listeyle karşılaştırma sonucunu gösterir.

    Modal DEĞİL: pencere açıkken ana penceredeki tablo hâlâ kullanılabilir,
    satırlar seçilip kimliklere bakılabilir. Bir satıra çift tıklanınca ana
    tablodaki karşılığı seçilip önizlemesi gösterilir.

    En önemli kategori "Bizde yok": dış listede kayıtlı olduğu halde bizim
    çıkardığımız sonuçlarda karşılığı bulunmayan kimlikler — asıl veri kaybı
    göstergesi."""

    RENKLER = {
        "Bizde yok": "#ff6b6b",
        "Listede yok": "#93c5fd",
        "Eşleşti ama veri düzeltilmesi gerek": "#c084fc",
        "Eşleşti": "#86efac",
    }
    ONCELIK = {"Bizde yok": 0, "Listede yok": 1,
              "Eşleşti ama veri düzeltilmesi gerek": 2, "Eşleşti": 3}

    def __init__(self, sonuc, bilgi, dosya_adi, ebeveyn=None):
        super().__init__(ebeveyn)
        self.setWindowTitle("Liste karşılaştırması")
        self.resize(1080, 640)
        # Modal olmasın: kullanıcı bu pencere açıkken ana tabloda gezinebilsin.
        self.setModal(False)

        self.ebeveyn_pencere = ebeveyn
        self.sonuc = sonuc
        self.satirlar = karsilastirma.satirlara_don(sonuc)
        self.gosterilen_satirlar = []
        self.aktif_filtre = None   # None = Tümü

        duzen = QVBoxLayout(self)

        baslik = QLabel(f"“{dosya_adi}” ile karşılaştırma")
        baslik.setStyleSheet("color: #f8fafc; font-size: 17px; font-weight: 700;")
        duzen.addWidget(baslik)

        ipucu = QLabel(
            "Bu pencere açıkken arkadaki listede gezinip kimliklere bakabilirsiniz. "
            "Aşağıdaki bir satıra çift tıklayınca o kimlik ana tabloda seçilir."
        )
        ipucu.setWordWrap(True)
        ipucu.setStyleSheet("color: #9aa3ad; font-size: 12px;")
        duzen.addWidget(ipucu)

        # --- Kategori kutucukları: sayı + renk, tıklanınca o kategoriye filtreler ---
        self.sayilar = Counter(s["Durum"] for s in self.satirlar)
        kutu_bar = QHBoxLayout()
        kutu_bar.setSpacing(8)
        self.kategori_grubu = QButtonGroup(self)
        self.kategori_grubu.setExclusive(True)
        self.kategori_dugmeleri = {}

        for durum in ("Bizde yok", "Listede yok", "Eşleşti ama veri düzeltilmesi gerek", "Eşleşti"):
            self._kategori_dugmesi_ekle(kutu_bar, durum, durum, self.sayilar.get(durum, 0),
                                        self.RENKLER[durum])
        self._kategori_dugmesi_ekle(kutu_bar, None, "Tümü", len(self.satirlar), "#e5e7eb")
        kutu_bar.addStretch(1)
        duzen.addLayout(kutu_bar)

        sutun_yazi = ", ".join(f"{ad}={sutun}" for ad, sutun in
                               sorted((bilgi.get("sutunlar") or {}).items()))
        ayrinti = QLabel(
            f"Listede okunan sayfa: {bilgi.get('sayfa', '-')}   •   "
            f"başlık satırı: {bilgi.get('baslik_satiri') or 'bulunamadı'}   •   "
            f"sütunlar: {sutun_yazi or '-'}"
        )
        ayrinti.setWordWrap(True)
        ayrinti.setStyleSheet("color: #6b7280; font-size: 11px;")
        duzen.addWidget(ayrinti)

        self.tablo = QTableWidget()
        self.basliklar = ["Durum", "Kimlik No", "Listedeki Ad Soyad",
                          "Bizdeki Ad Soyad", "Not"]
        self.tablo.setColumnCount(len(self.basliklar))
        self.tablo.setHorizontalHeaderLabels(self.basliklar)
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tablo.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tablo.setAlternatingRowColors(True)
        self.tablo.setWordWrap(False)
        self.tablo.verticalHeader().setVisible(False)
        for i, genislik in enumerate([120, 190, 220, 220, 300]):
            self.tablo.setColumnWidth(i, genislik)
        self.tablo.cellDoubleClicked.connect(self.satira_git)
        duzen.addWidget(self.tablo, 1)

        alt_ipucu = QLabel("")
        alt_ipucu.setStyleSheet("color: #6b7280; font-size: 11px;")
        self.satir_ipucu = alt_ipucu
        duzen.addWidget(alt_ipucu)

        dugmeler = QHBoxLayout()
        self.aktar_btn = QPushButton("Excel’e aktar ve aç")
        self.kapat_btn = QPushButton("Kapat")
        for b in (self.aktar_btn, self.kapat_btn):
            b.setMinimumHeight(36)
        dugmeler.addStretch(1)
        dugmeler.addWidget(self.aktar_btn)
        dugmeler.addWidget(self.kapat_btn)
        duzen.addLayout(dugmeler)

        self.aktar_btn.clicked.connect(self.excele_aktar)
        self.kapat_btn.clicked.connect(self.close)

        # Varsayılan olarak en önemli soruna odaklan: bizde olmayanlar varsa
        # onlar, yoksa sırayla diğer sorun kategorileri, hiçbiri yoksa Tümü.
        varsayilan = next(
            (d for d in ("Bizde yok", "Listede yok", "Eşleşti ama veri düzeltilmesi gerek")
             if self.sayilar.get(d)),
            None,
        )
        self.kategori_dugmeleri[varsayilan].setChecked(True)
        self.filtreyi_uygula(varsayilan)

    def _kategori_dugmesi_ekle(self, duzen, anahtar, etiket, sayi, renk):
        btn = QPushButton(f"{etiket}  ({sayi})")
        btn.setCheckable(True)
        btn.setMinimumHeight(34)
        btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #20242a; border: 2px solid {renk}66;
                border-radius: 8px; padding: 4px 12px; color: {renk}; font-weight: 700;
            }}
            QPushButton:hover {{ border-color: {renk}; }}
            QPushButton:checked {{ background-color: {renk}26; border-color: {renk}; }}
        """)
        btn.clicked.connect(lambda: self.filtreyi_uygula(anahtar))
        self.kategori_grubu.addButton(btn)
        self.kategori_dugmeleri[anahtar] = btn
        duzen.addWidget(btn)

    def filtreyi_uygula(self, durum):
        self.aktif_filtre = durum
        self.tabloyu_doldur()

    def tabloyu_doldur(self):
        satirlar = self.satirlar if self.aktif_filtre is None else \
            [s for s in self.satirlar if s["Durum"] == self.aktif_filtre]
        satirlar = sorted(satirlar, key=lambda s: self.ONCELIK.get(s["Durum"], 9))
        self.gosterilen_satirlar = satirlar

        self.tablo.setRowCount(0)
        for satir in satirlar:
            r = self.tablo.rowCount()
            self.tablo.insertRow(r)
            for c, baslik in enumerate(self.basliklar):
                item = QTableWidgetItem(str(satir.get(baslik, "")))
                if c == 0:
                    item.setForeground(QColor(self.RENKLER.get(satir["Durum"], "#f3f4f6")))
                self.tablo.setItem(r, c, item)

        if satirlar and any(s.get("_bizim_satir") is not None for s in satirlar):
            self.satir_ipucu.setText("Çift tıklayarak ilgili kimliği ana tabloda açabilirsiniz.")
        else:
            self.satir_ipucu.setText("")

    def satira_git(self, row, _col):
        if row < 0 or row >= len(self.gosterilen_satirlar):
            return
        bizim_satir = self.gosterilen_satirlar[row].get("_bizim_satir")
        if bizim_satir is None or self.ebeveyn_pencere is None:
            return
        goster = getattr(self.ebeveyn_pencere, "satiri_sec_ve_goster", None)
        if goster:
            goster(bizim_satir)

    def excele_aktar(self):
        if not self.satirlar:
            return
        try:
            fd, yol = tempfile.mkstemp(prefix="karsilastirma_", suffix=".xlsx")
            os.close(fd)

            disa_aktarilacak = [{k: v for k, v in s.items() if k in self.basliklar}
                                for s in self.satirlar]
            df = pd.DataFrame(disa_aktarilacak, columns=self.basliklar)
            df.to_excel(yol, index=False, engine="openpyxl")

            wb = load_workbook(yol)
            ws = wb.active
            ws.title = "Karşılaştırma"
            for hucre in ws[1]:
                hucre.font = Font(bold=True)

            dolgular = {
                "Bizde yok": PatternFill("solid", fgColor="FFC7CE"),
                "Listede yok": PatternFill("solid", fgColor="DDEBF7"),
                "Eşleşti ama veri düzeltilmesi gerek": PatternFill("solid", fgColor="E9D5FF"),
            }
            for i, satir in enumerate(self.satirlar, start=2):
                dolgu = dolgular.get(satir["Durum"])
                if dolgu:
                    for c in range(1, len(self.basliklar) + 1):
                        ws.cell(i, c).fill = dolgu

            for harf, genislik in zip("ABCDE", (14, 22, 28, 28, 40)):
                ws.column_dimensions[harf].width = genislik

            wb.save(yol)
            dosyayi_sistemde_ac(yol)
        except Exception as e:
            QMessageBox.critical(self, "Aktarma hatası", str(e))


class GecmisPenceresi(QDialog):
    """Kayıtlı taramaları listeler; seçileni açar, siler veya hepsini temizler."""

    def __init__(self, ebeveyn=None):
        super().__init__(ebeveyn)
        self.setWindowTitle("Geçmiş taramalar")
        self.resize(760, 460)
        self.secilen_yol = None

        duzen = QVBoxLayout(self)

        self.bilgi = QLabel("")
        self.bilgi.setWordWrap(True)
        duzen.addWidget(self.bilgi)

        self.liste = QTableWidget()
        self.basliklar = ["Tarih", "Dosyalar", "Kimlik", "Satır", "Boyut"]
        self.liste.setColumnCount(len(self.basliklar))
        self.liste.setHorizontalHeaderLabels(self.basliklar)
        self.liste.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.liste.setSelectionMode(QAbstractItemView.SingleSelection)
        self.liste.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.liste.setAlternatingRowColors(True)
        self.liste.verticalHeader().setVisible(False)
        for i, genislik in enumerate([160, 330, 80, 70, 90]):
            self.liste.setColumnWidth(i, genislik)
        self.liste.doubleClicked.connect(self.ac)
        duzen.addWidget(self.liste, 1)

        dugmeler = QHBoxLayout()
        self.ac_btn = QPushButton("Aç")
        self.sil_btn = QPushButton("Seçileni sil")
        self.temizle_btn = QPushButton("Geçmişi temizle")
        self.kapat_btn = QPushButton("Kapat")
        for b in (self.ac_btn, self.sil_btn, self.temizle_btn, self.kapat_btn):
            b.setMinimumHeight(36)
        self.temizle_btn.setStyleSheet(
            "QPushButton { background-color: #4a2020; border-color: #7a3030; }"
        )

        dugmeler.addWidget(self.ac_btn)
        dugmeler.addWidget(self.sil_btn)
        dugmeler.addStretch(1)
        dugmeler.addWidget(self.temizle_btn)
        dugmeler.addWidget(self.kapat_btn)
        duzen.addLayout(dugmeler)

        self.ac_btn.clicked.connect(self.ac)
        self.sil_btn.clicked.connect(self.sil)
        self.temizle_btn.clicked.connect(self.temizle)
        self.kapat_btn.clicked.connect(self.reject)

        self.yenile()

    def yenile(self):
        self.kayitlar = gecmis.taramalari_listele()
        self.liste.setRowCount(0)

        for kayit in self.kayitlar:
            r = self.liste.rowCount()
            self.liste.insertRow(r)
            try:
                tarih = datetime.fromisoformat(kayit.get("tarih", "")).strftime("%d.%m.%Y %H:%M")
            except (ValueError, TypeError):
                tarih = kayit.get("tarih", "")
            if kayit.get("guncellendi"):
                try:
                    tarih += datetime.fromisoformat(kayit["guncellendi"]).strftime(
                        "  (düzenlendi %H:%M)")
                except (ValueError, TypeError):
                    tarih += "  (düzenlendi)"
            dosyalar = ", ".join(kayit.get("dosyalar", []))
            degerler = [
                tarih,
                dosyalar,
                str(kayit.get("kimlik_sayisi", "")),
                str(kayit.get("satir_sayisi", "")),
                f"{kayit.get('boyut', 0) / 1024 / 1024:.1f} MB",
            ]
            for c, deger in enumerate(degerler):
                item = QTableWidgetItem(deger)
                if c == 1:
                    item.setToolTip(dosyalar)
                self.liste.setItem(r, c, item)

        toplam = sum(k.get("boyut", 0) for k in self.kayitlar)
        self.bilgi.setText(
            f"{len(self.kayitlar)} tarama saklanıyor (toplam {toplam / 1024 / 1024:.1f} MB). "
            f"En yeni {gecmis.AZAMI_KAYIT} tarama tutulur, eskiler kendiliğinden silinir.\n"
            f"Klasör: {gecmis.gecmis_dizini()}"
        )
        self.ac_btn.setEnabled(bool(self.kayitlar))
        self.sil_btn.setEnabled(bool(self.kayitlar))
        self.temizle_btn.setEnabled(bool(self.kayitlar))
        if self.kayitlar:
            self.liste.selectRow(0)

    def secili_kayit(self):
        r = self.liste.currentRow()
        if 0 <= r < len(self.kayitlar):
            return self.kayitlar[r]
        return None

    def ac(self):
        kayit = self.secili_kayit()
        if kayit:
            self.secilen_yol = kayit["yol"]
            self.accept()

    def sil(self):
        kayit = self.secili_kayit()
        if not kayit:
            return
        if QMessageBox.question(
            self, "Kaydı sil",
            f"{', '.join(kayit.get('dosyalar', []))} taraması silinsin mi?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        ) == QMessageBox.Yes:
            gecmis.taramayi_sil(kayit["yol"])
            self.yenile()

    def temizle(self):
        if QMessageBox.question(
            self, "Geçmişi temizle",
            "Kayıtlı bütün taramalar silinsin mi?\n"
            "Bu dosyalar kimlik numarası ve isim içerir; silme geri alınamaz.",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No,
        ) == QMessageBox.Yes:
            gecmis.gecmisi_temizle()
            self.yenile()


