import streamlit as st
import cv2
import numpy as np
import pymupdf
import pandas as pd
import time

from io import BytesIO
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from goruntu_isleme import kart_tespit_et_ve_duzelt
from metin_ayiklama import bilgileri_cimbizla


# =========================================================
# PDF
# =========================================================

# Kimlik alanlari kesimden sonra sabit OCR boyutuna normalize ediliyor. 200 DPI
# A4 sayfayi gereksiz buyutup hem render hem de feature cikarma maliyetini
# artiriyordu; 170 DPI ince baskiyi korurken sayfa pikselini yaklasik %28 azaltir.
PDF_RENDER_DPI = 170
MAX_TOPLAM_SAYFA = 200
MAX_TOPLAM_YUKLEME_BAYT = 300 * 1024 * 1024
MAX_GORUNTU_PIKSEL = 40_000_000


def _pixmap_to_bgr(pix):
    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
    if pix.n == 4:
        return cv2.cvtColor(img, cv2.COLOR_RGBA2BGR)
    if pix.n == 3:
        return cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    if pix.n == 1:
        return cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
    return img


def _yuklenen_dosyayi_goruntulere_ayir(dosya):
    ad_lower = dosya.name.lower()
    dosya_baslangic = time.perf_counter()

    if ad_lower.endswith(".pdf"):
        doc = None
        try:
            doc = pymupdf.open(stream=dosya.getvalue(), filetype="pdf")
            if doc.page_count == 0:
                yield {
                    "dosya_adi": dosya.name, "gorunen_isim": dosya.name,
                    "sayfa_no": None, "pdf_mi": True, "resim": None,
                    "hata": "PDF hiç sayfa içermiyor.",
                    "hazirlama_suresi": time.perf_counter() - dosya_baslangic,
                }
                return

            for sayfa_no, sayfa in enumerate(doc, start=1):
                sayfa_baslangic = time.perf_counter()
                try:
                    tahmini_w = max(1, int(sayfa.rect.width * PDF_RENDER_DPI / 72))
                    tahmini_h = max(1, int(sayfa.rect.height * PDF_RENDER_DPI / 72))
                    if tahmini_w * tahmini_h > MAX_GORUNTU_PIKSEL:
                        raise ValueError(
                            f"Sayfa çözünürlüğü güvenli sınırı aşıyor ({tahmini_w}×{tahmini_h})."
                        )
                    pix = sayfa.get_pixmap(dpi=PDF_RENDER_DPI)
                    resim = _pixmap_to_bgr(pix)
                    hata = None
                except Exception as e:
                    resim = None
                    hata = f"PDF sayfası görüntüye çevrilemedi: {e}"

                yield {
                    "dosya_adi": dosya.name, "gorunen_isim": f"{dosya.name} - Sayfa {sayfa_no}",
                    "sayfa_no": sayfa_no, "pdf_mi": True, "resim": resim, "hata": hata,
                    "hazirlama_suresi": time.perf_counter() - sayfa_baslangic,
                }
        except Exception as e:
            yield {
                "dosya_adi": dosya.name, "gorunen_isim": dosya.name,
                "sayfa_no": None, "pdf_mi": True, "resim": None,
                "hata": f"PDF açılamadı: {e}",
                "hazirlama_suresi": time.perf_counter() - dosya_baslangic,
            }
        finally:
            if doc is not None:
                doc.close()
        return

    try:
        ham_veri = dosya.getvalue()
        with Image.open(BytesIO(ham_veri)) as pil_resim:
            genislik, yukseklik = pil_resim.size
        if genislik <= 0 or yukseklik <= 0 or genislik * yukseklik > MAX_GORUNTU_PIKSEL:
            raise ValueError(f"Görüntü çözünürlüğü güvenli sınırı aşıyor ({genislik}×{yukseklik}).")

        veri = np.frombuffer(ham_veri, dtype=np.uint8)
        resim = cv2.imdecode(veri, cv2.IMREAD_COLOR)
        hata = None if resim is not None else "Görüntü dosyası okunamadı."
    except Exception as e:
        resim = None
        hata = f"Görüntü dosyası okunamadı: {e}"

    yield {
        "dosya_adi": dosya.name, "gorunen_isim": dosya.name,
        "sayfa_no": None, "pdf_mi": False, "resim": resim, "hata": hata,
        "hazirlama_suresi": time.perf_counter() - dosya_baslangic,
    }


def _yuklenen_dosya_is_sayisi(dosya):
    if not dosya.name.lower().endswith(".pdf"):
        return 1

    try:
        with pymupdf.open(stream=dosya.getvalue(), filetype="pdf") as doc:
            return max(1, doc.page_count)
    except Exception:
        # Hatalı dosya da sonuç tablosunda tek bir hata satırı olarak gösterilir.
        return 1


def _tum_yuklemeleri_goruntulere_ayir(dosyalar):
    for dosya in dosyalar:
        yield from _yuklenen_dosyayi_goruntulere_ayir(dosya)


# =========================================================
# SAYFA
# =========================================================

st.set_page_config(page_title="Kimlik Okuyucu", page_icon="🪪", layout="wide")
st.title("🪪 Kimlik Kartı Okuyucu")
st.caption("Yeni T.C. kimlik, eski nüfus cüzdanı ve göçmen/yabancı kimlik belgelerini otomatik olarak tespit eder.")


# =========================================================
# SESSION
# =========================================================

for anahtar, varsayilan in [
    ("sonuc_df", None), ("excel_data", None), ("bulunamayanlar", []),
    ("kimlik_pdf_data", None), ("kimlik_pdf_sayisi", 0),
    ("kimlik_pdf_duzeltilen", 0), ("kimlik_pdf_orijinal", 0),
]:
    if anahtar not in st.session_state:
        st.session_state[anahtar] = varsayilan


def sonuclari_temizle():
    """Yüklenen dosya listesi değişince önceki dosyalara ait çıktıları gizler."""
    st.session_state.sonuc_df = None
    st.session_state.excel_data = None
    st.session_state.bulunamayanlar = []
    st.session_state.kimlik_pdf_data = None
    st.session_state.kimlik_pdf_sayisi = 0
    st.session_state.kimlik_pdf_duzeltilen = 0
    st.session_state.kimlik_pdf_orijinal = 0


# =========================================================
# SIDEBAR
# =========================================================

with st.sidebar:
    st.header("Ayarlar")
    kimlik_pdf_olustur = st.checkbox(
        "📄 Kimlikleri PDF olarak birleştir", value=False,
        help="Kart bulunursa düzeltilmiş hali, bulunamazsa orijinal görüntü PDF'e eklenir.",
        on_change=sonuclari_temizle,
    )

    st.divider()
    debug_modu = st.checkbox("OCR debug", value=False)
    kart_siniri_goster = st.checkbox("Kart sınırını göster", value=False)
    ham_ocr_goster = st.checkbox("Ham OCR sonuçlarını göster", value=False)

debug_aktif = any([debug_modu, kart_siniri_goster, ham_ocr_goster])


# =========================================================
# DOSYA
# =========================================================

dosyalar = st.file_uploader(
    "Kimlik görsellerini veya PDF dosyalarını yükle",
    type=["jpg", "jpeg", "png", "pdf"],
    accept_multiple_files=True,
    max_upload_size=150,
    on_change=sonuclari_temizle,
)


# =========================================================
# ÖNİZLEME
# =========================================================

def onizleme_hazirla(resim):
    if resim is None:
        return None

    h, w = resim.shape[:2]
    if max(h, w) > 1000:
        oran = 1000 / max(h, w)
        resim = cv2.resize(resim, None, fx=oran, fy=oran, interpolation=cv2.INTER_AREA)

    basarili, encoded = cv2.imencode(".jpg", resim, [cv2.IMWRITE_JPEG_QUALITY, 82])
    return encoded.tobytes() if basarili else None


# =========================================================
# PDF'E RESİM EKLE
# =========================================================

def resmi_pdfe_ekle(pdf_doc, resim):
    if pdf_doc is None or resim is None:
        return False

    eklenen_sayfa_no = None
    try:
        h, w = resim.shape[:2]
        if h <= 0 or w <= 0:
            return False

        basarili, encoded = cv2.imencode(".jpg", resim, [cv2.IMWRITE_JPEG_QUALITY, 92])
        if not basarili:
            return False

        pdf_genislik = 720.0
        pdf_yukseklik = pdf_genislik * h / w
        eklenen_sayfa_no = pdf_doc.page_count
        sayfa = pdf_doc.new_page(width=pdf_genislik, height=pdf_yukseklik)
        rect = pymupdf.Rect(0, 0, pdf_genislik, pdf_yukseklik)
        sayfa.insert_image(rect, stream=encoded.tobytes(), keep_proportion=True)
        return True
    except Exception:
        # Sayfa oluşturulup resim ekleme başarısız olduysa boş sayfayı bırakma.
        if eklenen_sayfa_no is not None:
            try:
                if pdf_doc.page_count > eklenen_sayfa_no:
                    pdf_doc.delete_page(eklenen_sayfa_no)
            except Exception:
                pass
        return False


# =========================================================
# EXCEL
# =========================================================

def excel_olustur(df, meta):
    buffer = BytesIO()
    excel_df = df.copy()
    for sutun in excel_df.columns:
        excel_df[sutun] = excel_df[sutun].map(
            lambda deger: f"'{deger}" if isinstance(deger, str) and deger.lstrip().startswith(("=", "+", "-", "@")) else deger
        )
    excel_df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active
    ws.title = "Kimlik Sonuçları"

    kirmizi = PatternFill("solid", fgColor="FFC7CE")
    sari = PatternFill("solid", fgColor="FFF2CC")

    for hucre in ws[1]:
        hucre.font = Font(bold=True)

    basliklar = {hucre.value: hucre.column for hucre in ws[1]}
    kimlik_no_col, ad_col, soyad_col = basliklar["Kimlik No"], basliklar["Ad"], basliklar["Soyad"]
    bitis_tarihi_col = basliklar.get("Bitiş Tarihi")
    durum_col = basliklar.get("İşlem Durumu")

    for satir_no, item in enumerate(meta, start=2):
        # Süresi geçmiş göçmen belgesi satırını önce işaretle; eksik/düşük
        # güvenli hücrelerin kırmızı/sarı uyarıları bunun üzerine yazılsın.
        if item.get("belge_gecerli") is False:
            for sutun_no in range(1, ws.max_column + 1):
                hucre = ws.cell(satir_no, sutun_no)
                hucre.fill = sari
                hucre.font = Font(bold=True)

        if not item["kimlik_no_bulundu"]:
            ws.cell(satir_no, kimlik_no_col).fill = kirmizi

        if not item["ad_bulundu"]:
            ws.cell(satir_no, ad_col).fill = kirmizi
        elif item["ad_conf"] < 0.50:
            ws.cell(satir_no, ad_col).fill = kirmizi
        elif item["ad_conf"] < 0.70:
            ws.cell(satir_no, ad_col).fill = sari

        if not item["soyad_bulundu"]:
            ws.cell(satir_no, soyad_col).fill = kirmizi
        elif item["soyad_conf"] < 0.50:
            ws.cell(satir_no, soyad_col).fill = kirmizi
        elif item["soyad_conf"] < 0.70:
            ws.cell(satir_no, soyad_col).fill = sari

        if (
            bitis_tarihi_col is not None
            and item.get("belge_tipi") == "gocmen"
            and not item.get("bitis_tarihi_bulundu", False)
        ):
            ws.cell(satir_no, bitis_tarihi_col).fill = kirmizi

        if durum_col is not None and not item.get("islem_basarili", False):
            ws.cell(satir_no, durum_col).fill = kirmizi

    genislikler = {
        "Dosya": 32, "Sayfa No": 12, "Belge Türü": 20, "Kimlik No": 18,
        "Ad": 24, "Soyad": 24, "Bitiş Tarihi": 18, "Geçerlilik": 18,
        "Hazırlama (sn)": 16, "Tespit (sn)": 14, "OCR (sn)": 12,
        "Toplam (sn)": 14, "İşlem Durumu": 48,
    }
    for baslik, sutun_no in basliklar.items():
        ws.column_dimensions[get_column_letter(sutun_no)].width = genislikler.get(baslik, 18)

    sonuc = BytesIO()
    wb.save(sonuc)
    sonuc.seek(0)
    return sonuc.getvalue()


# =========================================================
# İŞLEM
# =========================================================

if dosyalar and st.button("🚀 İşlemi Başlat", type="primary"):
    sonuclari_temizle()

    durum = st.empty()
    durum.write("Dosyalar hazırlanıyor...")

    tum_sonuclar, meta, bulunamayanlar = [], [], []
    toplam_yukleme = sum(
        int(dosya.size) if getattr(dosya, "size", None) is not None else len(dosya.getvalue())
        for dosya in dosyalar
    )
    toplam = sum(_yuklenen_dosya_is_sayisi(dosya) for dosya in dosyalar)

    if toplam_yukleme > MAX_TOPLAM_YUKLEME_BAYT:
        durum.error("Toplam yükleme boyutu 300 MB sınırını aşıyor.")
        st.stop()
    if toplam == 0:
        durum.error("İşlenecek dosya bulunamadı.")
        st.stop()
    if toplam > MAX_TOPLAM_SAYFA:
        durum.error(f"En fazla {MAX_TOPLAM_SAYFA} sayfa tek işlemde işlenebilir.")
        st.stop()

    kimlik_pdf_doc = pymupdf.open() if kimlik_pdf_olustur else None
    pdf_toplam = pdf_duzeltilen = pdf_orijinal = 0

    progress = st.progress(0)

    # ---- Her sayfa -----------------------------------------------------
    for index, bilgi in enumerate(_tum_yuklemeleri_goruntulere_ayir(dosyalar)):
        sayfa_islem_baslangici = time.perf_counter()
        sayfa_no, pdf_mi, resim = bilgi["sayfa_no"], bilgi["pdf_mi"], bilgi["resim"]
        durum.write(f"Kimlikler işleniyor... {index + 1}/{toplam}")

        kimlik_no, ad, soyad = "Bulunamadi", "Bulunamadi", "Bulunamadi"
        ad_conf = soyad_conf = 0.0
        belge_tipi = "bilinmiyor"
        bitis_tarihi = ""
        belge_gecerli = None
        kart, kart_sonuc, ocr_sonuc = None, {}, {}
        girdi_hatasi = bilgi.get("hata")
        kart_hatasi = None
        ocr_hatasi = None
        hazirlama_suresi = float(bilgi.get("hazirlama_suresi", 0.0) or 0.0)
        tespit_suresi = 0.0
        ocr_suresi = 0.0

        # ---- Kart + belge tipi -------------------------------------------
        if resim is not None:
            try:
                kart_sonuc = kart_tespit_et_ve_duzelt(resim, debug_kart=kart_siniri_goster)
            except Exception as e:
                kart_sonuc = {"basarili": False, "mesaj": str(e)}
                kart_hatasi = str(e)

            if not isinstance(kart_sonuc, dict):
                kart_sonuc = {"basarili": False, "mesaj": "Kart tespiti geçerli bir sonuç sözlüğü döndürmedi."}

            if kart_sonuc.get("basarili", False):
                kart = kart_sonuc.get("kart")
                belge_tipi = kart_sonuc.get("belge_tipi", "tc")
            else:
                # Eşleşme türü bulunmuş fakat güvenli kesim üretilememiş
                # olabilir. Bu durumda kullanıcıya "Bilinmiyor" göstermek
                # yerine eşleşen belge türünü koru; OCR yine çalıştırılmaz.
                eslesen_tip = kart_sonuc.get("belge_tipi")
                if eslesen_tip in {"tc", "eski_tc", "gocmen"}:
                    belge_tipi = eslesen_tip
                kart_hatasi = kart_sonuc.get("mesaj") or "Kart tespit edilemedi."
            tespit_suresi = float(kart_sonuc.get("tespit_suresi", 0.0) or 0.0)

        # ---- Kimlik PDF ---------------------------------------------------
        if kimlik_pdf_olustur:
            duzeltilmis_mi = bool(kart is not None and kart_sonuc.get("duzeltildi", False))
            pdf_resmi = kart if duzeltilmis_mi else resim

            if pdf_resmi is not None and resmi_pdfe_ekle(kimlik_pdf_doc, pdf_resmi):
                pdf_toplam += 1
                if duzeltilmis_mi:
                    pdf_duzeltilen += 1
                else:
                    pdf_orijinal += 1

        # ---- OCR -----------------------------------------------------------
        sonuc_sayfa_no = sayfa_no
        if kart is not None:
            try:
                ocr_sonuc = bilgileri_cimbizla(
                    kart, sayfa_no=sayfa_no, debug=debug_modu or ham_ocr_goster, belge_tipi=belge_tipi
                )
            except Exception as e:
                ocr_sonuc = {"hata": str(e)}
                ocr_hatasi = str(e)

            if not isinstance(ocr_sonuc, dict):
                ocr_sonuc = {"hata": "OCR geçerli bir sonuç sözlüğü döndürmedi."}
                ocr_hatasi = ocr_sonuc["hata"]

            sonuc_sayfa_no = ocr_sonuc.get("sayfa_no", sayfa_no)
            kimlik_no = ocr_sonuc.get("tc_no", "Bulunamadi")
            ad = ocr_sonuc.get("ad", "Bulunamadi")
            soyad = ocr_sonuc.get("soyad", "Bulunamadi")
            try:
                ad_conf = float(ocr_sonuc.get("ad_conf", 0.0) or 0.0)
                soyad_conf = float(ocr_sonuc.get("soyad_conf", 0.0) or 0.0)
            except (TypeError, ValueError) as e:
                ad_conf = soyad_conf = 0.0
                ocr_hatasi = f"OCR güven değeri geçersiz: {e}"
            bitis_tarihi = ocr_sonuc.get("bitis_tarihi", "")
            belge_gecerli = ocr_sonuc.get("belge_gecerli")
            if ocr_sonuc.get("hata"):
                ocr_hatasi = str(ocr_sonuc["hata"])
            ocr_suresi = float(ocr_sonuc.get("ocr_suresi", 0.0) or 0.0)

        # ---- Belge türü adı / geçerlilik -----------------------------------
        belge_turu_yazi = {
            "gocmen": "Göçmen / Yabancı", "eski_tc": "Eski T.C. Kimlik", "tc": "T.C. Kimlik",
        }.get(belge_tipi, "Bilinmiyor")

        if belge_tipi == "gocmen":
            if belge_gecerli is False:
                gecerlilik_yazi = "GEÇERSİZ"
            elif belge_gecerli is True:
                gecerlilik_yazi = "Geçerli"
            else:
                gecerlilik_yazi = "Kontrol Edilemedi"
        else:
            gecerlilik_yazi = "-"
            bitis_tarihi = "-"

        # ---- Meta / işlem durumu --------------------------------------------
        kimlik_no_bulundu = kimlik_no != "Bulunamadi"
        ad_bulundu = ad != "Bulunamadi"
        soyad_bulundu = soyad != "Bulunamadi"
        bitis_tarihi_bulundu = belge_tipi != "gocmen" or bool(
            bitis_tarihi and bitis_tarihi not in {"-", "Bulunamadi"}
        )
        alan_durumlari = [
            ("Kimlik No", kimlik_no_bulundu),
            ("Ad", ad_bulundu),
            ("Soyad", soyad_bulundu),
        ]
        if belge_tipi == "gocmen":
            alan_durumlari.append(("Bitiş Tarihi", bitis_tarihi_bulundu))
        eksikler = [
            etiket for etiket, bulundu in
            alan_durumlari
            if not bulundu
        ]

        if girdi_hatasi:
            islem_durumu = f"Girdi hatası: {girdi_hatasi}"
        elif kart is None:
            islem_durumu = f"Tespit başarısız: {kart_hatasi or 'Kart bulunamadı.'}"
        elif ocr_hatasi:
            islem_durumu = f"OCR hatası: {ocr_hatasi}"
        elif eksikler:
            islem_durumu = f"Eksik alan: {', '.join(eksikler)}"
        else:
            islem_durumu = "Başarılı"

        toplam_sure = hazirlama_suresi + (time.perf_counter() - sayfa_islem_baslangici)

        # ---- Tablo -----------------------------------------------------------
        tum_sonuclar.append({
            "Dosya": bilgi["dosya_adi"], "Sayfa No": sonuc_sayfa_no if pdf_mi else "-",
            "Belge Türü": belge_turu_yazi, "Kimlik No": kimlik_no, "Ad": ad, "Soyad": soyad,
            "Bitiş Tarihi": bitis_tarihi, "Geçerlilik": gecerlilik_yazi,
            "Hazırlama (sn)": round(hazirlama_suresi, 3),
            "Tespit (sn)": round(tespit_suresi, 3), "OCR (sn)": round(ocr_suresi, 3),
            "Toplam (sn)": round(toplam_sure, 3),
            "İşlem Durumu": islem_durumu,
        })

        meta.append({
            "kimlik_no_bulundu": kimlik_no_bulundu, "ad_bulundu": ad_bulundu, "soyad_bulundu": soyad_bulundu,
            "ad_conf": ad_conf, "soyad_conf": soyad_conf, "belge_gecerli": belge_gecerli, "belge_tipi": belge_tipi,
            "bitis_tarihi_bulundu": bitis_tarihi_bulundu,
            "islem_basarili": islem_durumu == "Başarılı",
        })

        # ---- Bulunamayanlar -----------------------------------------------------
        inceleme_resmi = None
        if eksikler:
            # Kart tespit edilemediyse kullanicinin bulaniklik/kadraj sorununu
            # gorebilmesi icin orijinal sayfayi; OCR alani eksikse duzeltilmis
            # karti sakla. Ayni onizleme hem anlik debug kartinda hem de islem
            # sonundaki "Bulunamayan Kimlikler" bolumunde kullanilir.
            inceleme_resmi = onizleme_hazirla(
                kart if kart is not None else resim
            )
            bulunamayanlar.append({
                "pdf_mi": pdf_mi, "sayfa_no": sonuc_sayfa_no, "dosya": bilgi["dosya_adi"],
                "eksik": ", ".join(eksikler),
                "durum": islem_durumu,
                "resim": inceleme_resmi,
            })

        # ---- Debug -----------------------------------------------------------
        if debug_aktif:
            st.divider()
            st.write(f"### Sayfa {sayfa_no}" if pdf_mi else f"### {bilgi['dosya_adi']}")
            st.write(f"Belge tipi: **{belge_turu_yazi}**")
            st.write(f"Kimlik No: {kimlik_no}")
            st.write(f"Ad: {ad}")
            st.write(f"Soyad: {soyad}")
            st.write(f"İşlem durumu: {islem_durumu}")
            st.caption(
                f"Süre — hazırlama: {hazirlama_suresi:.3f} sn, "
                f"tespit: {tespit_suresi:.3f} sn, OCR: {ocr_suresi:.3f} sn, toplam: {toplam_sure:.3f} sn"
            )

            if kart is None and resim is not None:
                if inceleme_resmi is not None:
                    st.caption(
                        "İnceleme görüntüsü — kart tespit edilemediği için "
                        "orijinal sayfa gösteriliyor."
                    )
                    aday_hatalari = [
                        aday.get("hata")
                        for aday in kart_sonuc.get("aday_sonuclari", [])
                        if aday.get("hata")
                    ]
                    if aday_hatalari:
                        st.warning(f"Kesme nedeni: {aday_hatalari[0]}")
                    st.image(inceleme_resmi, width=650)
                else:
                    st.warning("İnceleme görüntüsü hazırlanamadı.")

            duzeltme_yontemi = kart_sonuc.get("duzeltme_yontemi")
            if duzeltme_yontemi == "kontur_on_tespit":
                st.caption(
                    "Göçmen kesme yolu: çerçeve ve tablo imzası doğrulandı; "
                    "yavaş referans/SIFT taraması atlandı."
                )
            elif duzeltme_yontemi == "kontur_tablo":
                st.caption(
                    "Göçmen kesme yolu: referans eşleşmesi belge tipini doğruladı; "
                    "köşeler tablo konturundan güvenli biçimde kurtarıldı."
                )
            elif duzeltme_yontemi == "affine_partial":
                st.caption(
                    "Göçmen kesme yolu: normal homografi reddedildi; "
                    "doğrulanmış hızlı affine yedeği kullanıldı."
                )

            if belge_tipi == "gocmen":
                st.write(f"Bitiş tarihi: {bitis_tarihi}")
                if ocr_sonuc.get("hizli_yol_kullanildi"):
                    st.caption(
                        "Göçmen OCR yolu: hızlı değer hücreleri; "
                        f"süre: {float(ocr_sonuc.get('hizli_hucre_ocr_suresi', 0.0) or 0.0):.3f} sn"
                    )
                elif ocr_sonuc.get("detector_fallback_kullanildi"):
                    st.caption(
                        "Göçmen OCR yolu: geometri/alan güveni nedeniyle detector fallback; "
                        f"hızlı deneme: {float(ocr_sonuc.get('hizli_hucre_ocr_suresi', 0.0) or 0.0):.3f} sn, "
                        f"detector: {float(ocr_sonuc.get('detector_ocr_suresi', 0.0) or 0.0):.3f} sn"
                    )
                fallback_denenen = ocr_sonuc.get("fallback_denenen_alanlar", [])
                fallback_kullanilan = ocr_sonuc.get("fallback_kullanilan_alanlar", [])
                if fallback_denenen:
                    st.caption(
                        "Alan hücresi fallback — denenen: "
                        f"{', '.join(fallback_denenen)}; kullanılan: "
                        f"{', '.join(fallback_kullanilan) if fallback_kullanilan else 'yok'}; "
                        f"süre: {float(ocr_sonuc.get('fallback_ocr_suresi', 0.0) or 0.0):.3f} sn"
                    )
                if belge_gecerli is False:
                    st.warning("⚠️ Bu belgenin geçerlilik süresi geçmiş.")
                elif belge_gecerli is True:
                    st.success("✅ Belge geçerli.")

            if debug_modu and ocr_sonuc.get("debug_resmi") is not None:
                st.image(cv2.cvtColor(ocr_sonuc["debug_resmi"], cv2.COLOR_BGR2RGB))

            if kart_siniri_goster and kart_sonuc.get("debug_resmi") is not None:
                st.image(cv2.cvtColor(kart_sonuc["debug_resmi"], cv2.COLOR_BGR2RGB))

            if debug_modu or kart_siniri_goster:
                with st.expander("Belge tespit ayrıntıları"):
                    st.json({
                        "secilen_tip": kart_sonuc.get("belge_tipi"),
                        "aday_sirasi": kart_sonuc.get("aday_sirasi"),
                        "duzeltildi": kart_sonuc.get("duzeltildi", False),
                        "duzeltme_yontemi": kart_sonuc.get("duzeltme_yontemi"),
                        "fallback": kart_sonuc.get("fallback", False),
                        "iyi_eslesme": kart_sonuc.get("iyi_eslesme", 0),
                        "inlier": kart_sonuc.get("inlier", 0),
                        "inlier_orani": kart_sonuc.get("inlier_orani", 0.0),
                        "on_tespit": kart_sonuc.get("on_tespit_metrikleri", {}),
                        "mesaj": kart_sonuc.get("mesaj", ""),
                        "adaylar": kart_sonuc.get("aday_sonuclari", []),
                        "referans_hatalari": kart_sonuc.get("referans_hatalari", {}),
                    })

            if ham_ocr_goster:
                with st.expander("Ham OCR"):
                    ham_sonuclar = ocr_sonuc.get("tum_ocr", [])
                    if not ham_sonuclar:
                        st.info("OCR sonucu yok.")
                    for item in ham_sonuclar:
                        st.write(f'{item["text"]} — {item["conf"]}')

        progress.progress((index + 1) / toplam)

    # ---- PDF tamamla -----------------------------------------------------
    if kimlik_pdf_doc is not None:
        if pdf_toplam > 0:
            try:
                st.session_state.kimlik_pdf_data = kimlik_pdf_doc.tobytes(garbage=3, deflate=True)
                st.session_state.kimlik_pdf_sayisi = pdf_toplam
                st.session_state.kimlik_pdf_duzeltilen = pdf_duzeltilen
                st.session_state.kimlik_pdf_orijinal = pdf_orijinal
            except Exception:
                st.session_state.kimlik_pdf_data = None
        kimlik_pdf_doc.close()

    # ---- Sonuç -----------------------------------------------------------
    tam_basarili = sum(item["İşlem Durumu"] == "Başarılı" for item in tum_sonuclar)
    sorunlu = len(tum_sonuclar) - tam_basarili
    if sorunlu:
        durum.warning(f"{len(tum_sonuclar)} sayfa işlendi; {tam_basarili} başarılı, {sorunlu} kontrol edilmeli.")
    else:
        durum.success(f"✅ {len(tum_sonuclar)} sayfa başarıyla işlendi.")

    df = pd.DataFrame(
        tum_sonuclar,
        columns=[
            "Dosya", "Sayfa No", "Belge Türü", "Kimlik No", "Ad", "Soyad",
            "Bitiş Tarihi", "Geçerlilik", "Hazırlama (sn)", "Tespit (sn)",
            "OCR (sn)", "Toplam (sn)", "İşlem Durumu",
        ],
    )
    st.session_state.sonuc_df = df
    st.session_state.excel_data = excel_olustur(df, meta)
    st.session_state.bulunamayanlar = bulunamayanlar


# =========================================================
# SONUÇ TABLOSU
# =========================================================

if st.session_state.sonuc_df is not None:
    st.divider()
    st.header("📋 Sonuçlar")
    sonuc_df = st.session_state.sonuc_df
    hazirlama_toplam = float(sonuc_df["Hazırlama (sn)"].sum())
    tespit_toplam = float(sonuc_df["Tespit (sn)"].sum())
    ocr_toplam = float(sonuc_df["OCR (sn)"].sum())
    sayfa_toplam = float(sonuc_df["Toplam (sn)"].sum())
    diger_toplam = max(0.0, sayfa_toplam - hazirlama_toplam - tespit_toplam - ocr_toplam)
    st.caption(
        f"Süre dağılımı — hazırlama: {hazirlama_toplam:.2f} sn, "
        f"tespit: {tespit_toplam:.2f} sn, OCR: {ocr_toplam:.2f} sn, "
        f"diğer işlemler: {diger_toplam:.2f} sn. "
        "İlk OCR çağrısı EasyOCR modelinin belleğe yüklenmesini de içerir."
    )
    st.dataframe(sonuc_df, use_container_width=True, hide_index=True)

    if st.session_state.kimlik_pdf_data is not None:
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                "📥 Excel indir", data=st.session_state.excel_data, file_name="kimlik_sonuclari.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col2:
            st.download_button(
                f"📄 Kimlik PDF'ini indir ({st.session_state.kimlik_pdf_sayisi})",
                data=st.session_state.kimlik_pdf_data, file_name="kimlikler.pdf", mime="application/pdf",
                use_container_width=True,
            )
        st.caption(
            f"PDF: {st.session_state.kimlik_pdf_duzeltilen} belge düzeltilmiş, "
            f"{st.session_state.kimlik_pdf_orijinal} belge tespit edilemediği için orijinal haliyle eklendi."
        )
    else:
        st.download_button(
            "📥 Excel indir", data=st.session_state.excel_data, file_name="kimlik_sonuclari.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    bulunamayanlar = st.session_state.bulunamayanlar
    if bulunamayanlar:
        with st.expander(f"⚠️ Bulunamayan Kimlikler ({len(bulunamayanlar)})"):
            for item in bulunamayanlar:
                st.markdown("---")
                if item["pdf_mi"]:
                    st.subheader(f'📄 {item["dosya"]} — Sayfa {item["sayfa_no"]}')
                else:
                    st.subheader(f'🖼️ {item["dosya"]}')

                st.write(f'**Bulunamayan alanlar:** {item["eksik"]}')
                st.caption(item.get("durum", ""))
                if item["resim"] is not None:
                    st.image(item["resim"], width=650)
