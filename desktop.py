# -*- coding: utf-8 -*-
"""Kimlik Okuyucu masaustu uygulamasi -- giris noktasi.

Kod tabani asagidaki modullere bolundu (eskiden hepsi bu dosyadaydi, dosya
3000 satira yaklasinca okunmasi/degistirilmesi zorlasmisti):

    gorsel_araclari.py   PDF/Excel/JPG <-> OpenCV/Qt goruntu donusumleri
    worker.py             Worker, AlanOkuyucu (arka plan tespit/OCR is parcaciklari)
    pencereler.py          KarsilastirmaPenceresi, GecmisPenceresi
    onizleme_widget.py     OnizlemeEtiketi (fareyle bolge secilen onizleme)
    desktop_stil.py        koyu tema QSS'i
    gecmis.py, karsilastirma.py, excel_kaynak.py  (onceden de ayriydi)

Bu dosyada yalnizca MainWindow -- yani asil pencere/tablo/is akisi -- kaliyor."""

import sys
import os
import time
from datetime import datetime
import tempfile

import cv2
import numpy as np
import pandas as pd
import pymupdf
from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font

from PySide6.QtCore import Qt, QEvent, QTimer
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QFileDialog, QLabel, QTableWidget, QTableWidgetItem,
    QProgressBar, QMessageBox, QCheckBox, QSplitter, QHeaderView,
    QAbstractItemView, QTextEdit, QScrollArea, QSizePolicy, QFrame,
    QRubberBand, QDialog, QButtonGroup
)

from goruntu_isleme import sayfa_sirasina_diz
import gecmis
import excel_kaynak
import karsilastirma

from gorsel_araclari import (
    pixmap_to_bgr, dosyayi_goruntulere_ayir, resmi_pdfe_ekle, bgr_to_pixmap,
    dosyayi_sistemde_ac, PDF_RENDER_DPI, DESTEKLENEN,
)
from worker import Worker, AlanOkuyucu
from pencereler import KarsilastirmaPenceresi, GecmisPenceresi
from onizleme_widget import OnizlemeEtiketi
from desktop_stil import KOYU_TEMA

BUILD_ID = "DESKTOP-IMAGE-FIX-V10"
print(f"[DESKTOP BUILD] {BUILD_ID}")
print(f"[DESKTOP FILE] {__file__}")
print(f"[CV2 FILE] {getattr(cv2, '__file__', '?')}")
print(f"[CV2 VERSION] {getattr(cv2, '__version__', '?')}")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Kimlik Okuyucu — Masaüstü")
        self.resize(1280, 760)
        self.setMinimumSize(980, 620)

        self.dosyalar = []
        self.sonuclar = []
        self.canli_sonuclar = []
        self.pdf_bytes = None
        self.worker = None

        # Önizleme: zoom = None ise pencereye sığdırılır, sayı ise o oranda gösterilir.
        self.zoom = None
        self._sayfa_onbellek = {}
        self._isaretli_onbellek = (None, None)

        # Tablodaki sonuçların yazıldığı geçmiş kaydı. Düzenleme yapıldıkça
        # AYNI kayıt güncelleniyor; her düzenleme için yeni kayıt açılmıyor.
        self.gecmis_yolu = None
        self.gecmis_zamanlayici = QTimer(self)
        self.gecmis_zamanlayici.setSingleShot(True)
        self.gecmis_zamanlayici.setInterval(2500)
        self.gecmis_zamanlayici.timeout.connect(self.gecmisi_guncelle)

        # Fareyle seçilen alan ve o alanı okuyan iş parçacığı
        self._secim = None                # görüntü koordinatında (x0, y0, x1, y1)
        self._onizleme_ham = None         # işaretsiz görüntü (kırpma buradan alınır)
        self._onizleme_olcek = 1.0
        self._onizleme_ofset = (0, 0)     # gösterilen görüntünün sayfadaki yeri
        self.alan_okuyucu = None

        # "kart": satıra tıklanınca o kimliğe yakınlaşılır.
        # "sayfa": Sayfa sütununa tıklanınca sayfanın tamamı gösterilir.
        self._onizleme_kipi = "kart"

        self.edit_mode = False
        self._table_updating = False

        self.setStyleSheet(KOYU_TEMA)

        root = QWidget()
        self.setCentralWidget(root)
        ana = QVBoxLayout(root)

        ust = QHBoxLayout()
        ust.setSpacing(10)

        self.sec_btn = QPushButton("Dosya Seç")
        self.klasor_btn = QPushButton("Klasör Seç")
        self.baslat_btn = QPushButton("İşlemi Başlat")
        self.baslat_btn.setEnabled(False)

        self.karsilastir_btn = QPushButton("Liste Karşılaştır")
        self.karsilastir_btn.setToolTip(
            "Dışarıdan gelen kimlik listesini (Excel) yükleyip tablodaki sonuçlarla "
            "karşılaştırır: kaç kayıt tuttu, hangileri tutmadı, hangi kimlikler eksik."
        )
        self.karsilastir_btn.setEnabled(False)

        self.gecmis_btn = QPushButton("Geçmiş")
        self.gecmis_btn.setToolTip(
            "Yakın zamanlı taramalar burada saklanır; eskisini açıp Excel/PDF alabilirsiniz."
        )

        self.edit_btn = QPushButton("Düzenle")
        self.edit_btn.setCheckable(True)
        self.edit_btn.setEnabled(False)

        # Bu üç seçenek artık arayüzde gösterilmiyor — davranışları her zaman
        # açık kabul ediliyor (çok kimlikli sayfa taraması, derin okuma,
        # kurtarma). Widget'lar yine de oluşturuluyor (checked=True) ve
        # layout'a eklenmiyor: kod tabanının geri kalanı hâlâ
        # `self.X_cb.isChecked()` üzerinden okuyor, tek değişen bunların
        # artık kullanıcıya görünmemesi.
        self.coklu_cb = QCheckBox("Çok kimlikli sayfa")
        self.coklu_cb.setChecked(True)
        self.derin_cb = QCheckBox("Derin okuma")
        self.derin_cb.setChecked(True)
        self.kurtar_cb = QCheckBox("Kurtarma")
        self.kurtar_cb.setChecked(True)

        self.debug_cb = QCheckBox("Debug")
        self.debug_cb.setToolTip("Seçilen satırın belge tespit / OCR ayrıntılarını gösterir.")

        for btn in (
            self.sec_btn,
            self.klasor_btn,
            self.baslat_btn,
            self.gecmis_btn,
            self.karsilastir_btn,
            self.edit_btn,
        ):
            btn.setMinimumHeight(42)

        ust.addWidget(self.sec_btn)
        ust.addWidget(self.klasor_btn)
        ust.addWidget(self.baslat_btn)
        ust.addWidget(self.gecmis_btn)
        ust.addWidget(self.karsilastir_btn)
        ust.addWidget(self.edit_btn)
        ust.addStretch(1)
        ust.addWidget(self.debug_cb)
        ana.addLayout(ust)

        self.bilgi = QLabel("Dosya seçilmedi.")
        self.bilgi.setWordWrap(True)
        ana.addWidget(self.bilgi)

        # Okunamayan alanların sayısı buraya yazılır
        # ("3 Kimlik No bulunamadı, 1 Ad bulunamadı..."), satırlar
        # geldikçe canlı güncellenir.
        self.ozet = QLabel("")
        self.ozet.setWordWrap(True)
        self.ozet.setStyleSheet("color: #cbd5e1; font-size: 13px;")
        ana.addWidget(self.ozet)

        self.progress = QProgressBar()
        ana.addWidget(self.progress)

        # Kaydetme butonlarını ana içerikten bağımsız ve her zaman görünür tut.
        kaydet_bar = QHBoxLayout()
        kaydet_bar.setSpacing(10)
        kaydet_bar.addStretch(1)

        self.excel_btn = QPushButton("Excel’e Çevir ve Aç")
        self.pdf_btn = QPushButton("PDF’e Çevir ve Aç")
        self.excel_btn.setEnabled(False)
        self.pdf_btn.setEnabled(False)
        self.excel_btn.setMinimumHeight(42)
        self.pdf_btn.setMinimumHeight(42)

        kaydet_bar.addWidget(self.excel_btn)
        kaydet_bar.addWidget(self.pdf_btn)
        ana.addLayout(kaydet_bar)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setChildrenCollapsible(False)
        ana.addWidget(splitter, 1)

        self.table = QTableWidget()
        self.headers = [
            "Dosya", "Sayfa", "Kart", "Belge Türü", "Kimlik No", "Ad", "Soyad",
            "Bitiş Tarihi", "Geçerlilik", "Durum", "Süre",
        ]
        self.table.setColumnCount(len(self.headers))
        self.table.setHorizontalHeaderLabels(self.headers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setWordWrap(False)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.verticalHeader().setDefaultSectionSize(32)

        header = self.table.horizontalHeader()
        header.setSectionsMovable(True)
        header.setStretchLastSection(False)

        for idx in range(len(self.headers)):
            header.setSectionResizeMode(idx, QHeaderView.Interactive)

        genislikler = [190, 65, 60, 145, 145, 175, 175, 120, 115, 230, 70]
        for idx, width in enumerate(genislikler):
            self.table.setColumnWidth(idx, width)

        splitter.addWidget(self.table)

        sag = QWidget()
        saglay = QVBoxLayout(sag)
        saglay.setContentsMargins(6, 0, 0, 0)

        self.sag_splitter = QSplitter(Qt.Vertical)
        self.sag_splitter.setChildrenCollapsible(False)

        preview_frame = QFrame()
        preview_layout = QVBoxLayout(preview_frame)
        preview_layout.setContentsMargins(4, 4, 4, 4)

        # Sayfada kaç kimlik bulunduğu + zoom kontrolleri
        onizleme_bar = QHBoxLayout()
        onizleme_bar.setSpacing(6)

        # Tek satır kalsın: sarmalanırsa panelin yarısını kaplayıp görüntüye
        # yer bırakmıyordu. Sığmayan kısım "…" ile kısaltılıyor.
        self.sayfa_bilgi = QLabel("")
        self.sayfa_bilgi.setWordWrap(False)
        self.sayfa_bilgi.setMinimumWidth(80)
        self.sayfa_bilgi.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed)
        self.sayfa_bilgi.setStyleSheet("color: #dce1e7; font-size: 13px; font-weight: 600;")
        self._sayfa_bilgi_metni = ""

        self.alan_oku_btn = QPushButton("🔍 Seçili alanı oku")
        self.alan_oku_btn.setEnabled(False)
        self.alan_oku_btn.setToolTip(
            "İşaretlenmemiş bir kimliği fareyle çerçeveleyin, sonra bu düğmeye basın."
        )
        self.alan_oku_btn.setMinimumHeight(30)
        self.alan_oku_btn.setMaximumHeight(30)

        self.zoom_azalt_btn = QPushButton("−")
        self.zoom_arttir_btn = QPushButton("+")
        self.zoom_sigdir_btn = QPushButton("Sığdır")
        for btn in (self.zoom_azalt_btn, self.zoom_arttir_btn, self.zoom_sigdir_btn):
            btn.setMinimumHeight(30)
            btn.setMaximumHeight(30)
        self.zoom_azalt_btn.setMaximumWidth(38)
        self.zoom_arttir_btn.setMaximumWidth(38)
        self.zoom_sigdir_btn.setMaximumWidth(78)

        # Bilgi yazısı kendi satırında: düğmelerle aynı satırda sıkışıp
        # "Sayfa 2 — 3…" gibi kesiliyordu.
        preview_layout.addWidget(self.sayfa_bilgi)

        onizleme_bar.addWidget(self.alan_oku_btn)
        onizleme_bar.addStretch(1)
        onizleme_bar.addWidget(self.zoom_azalt_btn)
        onizleme_bar.addWidget(self.zoom_arttir_btn)
        onizleme_bar.addWidget(self.zoom_sigdir_btn)
        preview_layout.addLayout(onizleme_bar)

        self.preview = OnizlemeEtiketi("Bir satır seçince kimlik burada gösterilecek.")
        self.preview.setAlignment(Qt.AlignCenter)
        self.preview.setMinimumSize(320, 260)
        self.preview.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.preview.setStyleSheet(
            "border: 1px solid #454b53; background: #111316; color: #dce1e7; border-radius: 7px;"
        )

        self.preview_scroll = QScrollArea()
        self.preview_scroll.setWidgetResizable(True)
        self.preview_scroll.setFrameShape(QFrame.NoFrame)
        self.preview_scroll.setWidget(self.preview)
        self.preview_scroll.viewport().installEventFilter(self)
        preview_layout.addWidget(self.preview_scroll, 1)

        # Uzun açıklama artık başlıkta değil; görüntünün altında tek satır ipucu.
        self.onizleme_ipucu = QLabel("")
        self.onizleme_ipucu.setWordWrap(False)
        self.onizleme_ipucu.setMinimumWidth(80)
        self.onizleme_ipucu.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Fixed)
        self.onizleme_ipucu.setStyleSheet("color: #9aa3ad; font-size: 12px;")
        self._onizleme_ipucu_metni = ""
        preview_layout.addWidget(self.onizleme_ipucu)

        self.debug_text = QTextEdit()
        self.debug_text.setReadOnly(True)
        self.debug_text.setPlaceholderText(
            "Debug açıksa seçilen satırın belge tespit / OCR ayrıntıları burada görünür."
        )
        self.debug_text.setMinimumHeight(140)

        self.sag_splitter.addWidget(preview_frame)
        self.sag_splitter.addWidget(self.debug_text)

        # Debug kapalı başlar; böylece kimlik önizlemesi bütün sağ paneli kullanır.
        self.debug_text.hide()
        self.sag_splitter.setSizes([1, 0])

        saglay.addWidget(self.sag_splitter, 1)
        splitter.addWidget(sag)
        # Sağ panel artık sayfa görüntüsünü de gösteriyor; biraz daha geniş başlasın.
        splitter.setSizes([780, 540])

        self.sec_btn.clicked.connect(self.dosya_sec)
        self.klasor_btn.clicked.connect(self.klasor_sec)
        self.baslat_btn.clicked.connect(self.baslat)
        self.edit_btn.toggled.connect(self.duzenleme_modu_degisti)
        self.debug_cb.toggled.connect(self.debug_gorunumu_degisti)
        self.table.itemSelectionChanged.connect(self.onizleme_goster)
        self.table.cellClicked.connect(self.hucreye_tiklandi)
        self.table.cellChanged.connect(self.hucre_degisti)
        self.excel_btn.clicked.connect(self.excel_kaydet)
        self.pdf_btn.clicked.connect(self.pdf_kaydet)
        self.gecmis_btn.clicked.connect(self.gecmisi_goster)
        self.karsilastir_btn.clicked.connect(self.listeyle_karsilastir)
        self.zoom_arttir_btn.clicked.connect(lambda: self.zoom_degistir(1.25))
        self.zoom_azalt_btn.clicked.connect(lambda: self.zoom_degistir(0.8))
        self.zoom_sigdir_btn.clicked.connect(self.zoom_sigdir)
        self.coklu_cb.toggled.connect(lambda _: self.onizleme_goster())
        self.preview.alan_secildi.connect(self.alan_secildi)
        self.alan_oku_btn.clicked.connect(self.secili_alani_oku)

    def dosya_sec(self):
        yollar, _ = QFileDialog.getOpenFileNames(
            self, "Kimlik dosyalarını seç", "",
            "Kimlik / PDF / Excel (*.jpg *.jpeg *.png *.pdf *.xlsx *.xlsm)"
        )
        if yollar:
            self.dosyalar = yollar
            self.bilgi.setText(f"{len(yollar)} dosya seçildi.")
            self.baslat_btn.setEnabled(True)

    def klasor_sec(self):
        klasor = QFileDialog.getExistingDirectory(self, "Klasör seç")
        if not klasor:
            return
        yollar = []
        for kok, _, dosyalar in os.walk(klasor):
            for ad in dosyalar:
                yol = os.path.join(kok, ad)
                if os.path.splitext(ad)[1].lower() in DESTEKLENEN:
                    yollar.append(yol)
        self.dosyalar = sorted(yollar)
        self.bilgi.setText(f"{len(yollar)} dosya bulundu.")
        self.baslat_btn.setEnabled(bool(yollar))

    def baslat(self):
        if not self.dosyalar:
            return
        self.gecmisi_guncelle()      # önceki taramanın bekleyen düzenlemesi kaybolmasın

        self.table.setRowCount(0)
        self.sonuclar = []
        self.canli_sonuclar = []
        self.ozet.setText("")
        self.gecmis_yolu = None
        self.pdf_bytes = None
        self.excel_btn.setEnabled(False)
        self.pdf_btn.setEnabled(False)
        self.edit_btn.setChecked(False)
        self.edit_btn.setEnabled(False)
        self.baslat_btn.setEnabled(False)
        self.karsilastir_btn.setEnabled(False)
        self.progress.setValue(0)
        self.debug_text.clear()

        self._sayfa_onbellek = {}

        self.worker = Worker(
            self.dosyalar,
            self.debug_cb.isChecked(),
            self.kurtar_cb.isChecked(),
            self.coklu_cb.isChecked(),
            self.derin_cb.isChecked(),
        )
        self.worker.progress.connect(self.progress_guncelle)
        self.worker.row_ready.connect(self.satir_ekle)
        self.worker.finished_ok.connect(self.bitti)
        self.worker.failed.connect(self.hata)
        self.worker.start()

    def progress_guncelle(self, n, toplam, mesaj):
        self.progress.setMaximum(max(1, toplam))
        self.progress.setValue(n)
        self.bilgi.setText(f"{mesaj} — {n}/{toplam}")

    def satir_ekle(self, row):
        # Geçersiz belgelerde satırı sarıya boyamıyoruz (koyu temada okunmuyordu);
        # renk kuralları tabloya_satir_yaz içinde toplandı.
        self.canli_sonuclar.append(row)
        self._table_updating = True
        try:
            self.tabloya_satir_yaz(row)
        finally:
            self._table_updating = False

        self.ozet_guncelle(self.canli_sonuclar)


    def eksik_sayilari(self, satirlar):
        """Okunamayan alanları sayar: (alan -> adet, kartı hiç bulunamayan sayfa
        sayısı). Bitiş tarihi yalnız göçmen belgelerinde beklenir."""
        sayac = {alan: 0 for alan in ("Kimlik No", "Ad", "Soyad", "Bitiş Tarihi")}
        kart_bulunamayan = 0
        supheli = 0

        for r in satirlar:
            if not r.get("_kart_bulundu", True):
                kart_bulunamayan += 1
                continue
            if r.get("_tc_supheli"):
                supheli += 1
            for alan in ("Kimlik No", "Ad", "Soyad"):
                if str(r.get(alan, "")).strip() in ("", "-", "Bulunamadi"):
                    sayac[alan] += 1
            if r.get("_belge_tipi") == "gocmen":
                if str(r.get("Bitiş Tarihi", "")).strip() in ("", "-"):
                    sayac["Bitiş Tarihi"] += 1

        return sayac, kart_bulunamayan, supheli

    def ozet_guncelle(self, satirlar):
        """Bulunamayan alanların özetini bilgi çubuğunun altına yazar."""
        if not satirlar:
            self.ozet.setText("")
            return

        sayac, kart_bulunamayan, supheli = self.eksik_sayilari(satirlar)

        parcalar = []
        if kart_bulunamayan:
            parcalar.append(f"{kart_bulunamayan} sayfada kimlik bulunamadı")
        parcalar.extend(
            f"{adet} {alan} bulunamadı" for alan, adet in sayac.items() if adet
        )
        if supheli:
            parcalar.append(f"{supheli} Kimlik No doğrulanamadı")

        if parcalar:
            self.ozet.setText("Bulunamayanlar →  " + "   ·   ".join(parcalar))
            self.ozet.setStyleSheet("color: #f0b429; font-size: 13px; font-weight: 600;")
        else:
            okunan = len(satirlar) - kart_bulunamayan
            self.ozet.setText(
                f"Bulunamayan alan yok — {okunan} kimliğin tüm alanları okundu."
            )
            self.ozet.setStyleSheet("color: #86efac; font-size: 13px; font-weight: 600;")

    def debug_gorunumu_degisti(self, aktif):
        """
        Debug alanını yalnız checkbox açıkken göster.
        Kapalıyken sağ panelin tamamını kimlik önizlemesine bırak.
        """
        if aktif:
            self.debug_text.show()
            self.sag_splitter.setSizes([620, 240])
        else:
            self.debug_text.hide()
            self.sag_splitter.setSizes([1, 0])

        # Seçili kimliği yeni kullanılabilir alana yeniden sığdır.
        self.onizleme_goster()

    def duzenleme_modu_degisti(self, aktif):
        self.edit_mode = bool(aktif)

        if self.edit_mode:
            self.edit_btn.setText("Düzenlemeyi Bitir")
            self.table.setEditTriggers(
                QAbstractItemView.DoubleClicked
                | QAbstractItemView.SelectedClicked
                | QAbstractItemView.EditKeyPressed
            )
            self.bilgi.setText(
                "Düzenleme modu açık — Kimlik No, Ad ve Soyad hücrelerini değiştirebilirsin."
            )
        else:
            self.edit_btn.setText("Düzenle")
            self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        editable_headers = {"Kimlik No", "Ad", "Soyad"}

        self._table_updating = True
        try:
            for r in range(self.table.rowCount()):
                for c, baslik in enumerate(self.headers):
                    item = self.table.item(r, c)
                    if item is None:
                        continue

                    if self.edit_mode and baslik in editable_headers:
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                    else:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
        finally:
            self._table_updating = False

    def hucre_degisti(self, row, column):
        if self._table_updating:
            return

        if row < 0 or row >= len(self.sonuclar):
            return

        baslik = self.headers[column]
        if baslik not in {"Kimlik No", "Ad", "Soyad"}:
            return

        item = self.table.item(row, column)
        if item is None:
            return

        yeni_deger = item.text().strip() or "Bulunamadi"

        # Ana veri modelini de güncelle:
        # Excel kaydı self.sonuclar'dan üretildiği için değişiklik doğrudan yansır.
        self.sonuclar[row][baslik] = yeni_deger
        self.sonuclar[row]["_manuel_duzenlendi"] = True
        # Hangi alanların elle değiştirildiği saklanıyor ki tablo yeniden
        # kurulduğunda (sıralama değişince) vurgu kaybolmasın.
        duzenlenenler = set(self.sonuclar[row].get("_duzenlenen_alanlar") or [])
        duzenlenenler.add(baslik)
        self.sonuclar[row]["_duzenlenen_alanlar"] = sorted(duzenlenenler)
        # Durum yazısı OCR sonucunu anlatıyordu; elle doldurulan alandan sonra
        # "Eksik alan: Ad" gibi yanıltıcı kalıyordu. Yeniden hesaplanıyor.
        satir = self.sonuclar[row]
        satir["_tc_supheli"] = False
        eksikler = [
            alan for alan in ("Kimlik No", "Ad", "Soyad")
            if str(satir.get(alan, "")).strip() in ("", "Bulunamadi")
        ]
        onek = "Elle eklendi" if satir.get("_manuel_eklendi") else "Elle düzeltildi"
        satir["Durum"] = onek if not eksikler else f"{onek} — eksik: " + ", ".join(eksikler)

        self._table_updating = True
        try:
            durum_item = self.table.item(row, self.headers.index("Durum"))
            if durum_item is not None:
                durum_item.setText(satir["Durum"])
                durum_item.setForeground(QColor("#93c5fd"))
        finally:
            self._table_updating = False

        self.duzenlenmis_hucreyi_boya(item)
        self.ozet_guncelle(self.sonuclar)
        self.gecmisi_guncellemeyi_planla()

        if self.table.currentRow() == row:
            self.onizleme_goster()

    def bitti(self, sonuclar, pdf_bytes, durduruldu=False):
        self.sonuclar = sonuclar
        self.pdf_bytes = pdf_bytes
        self.baslat_btn.setEnabled(True)
        self.excel_btn.setEnabled(bool(sonuclar))
        self.pdf_btn.setEnabled(bool(sonuclar))
        self.edit_btn.setEnabled(bool(sonuclar))
        self.karsilastir_btn.setEnabled(bool(sonuclar))
        self.canli_sonuclar = list(sonuclar)

        # Bir sayfada birden fazla kimlik olabildiği için satır sayısı ile
        # sayfa sayısı artık aynı değil; ikisini de ayrı ayrı gösteriyoruz.
        kimlik_sayisi = sum(1 for r in sonuclar if r.get("_kart_bulundu"))
        kurtarilan = sum(1 for r in sonuclar if r.get("_kurtarildi"))
        sayfa_sayisi = len({(r.get("Dosya"), r.get("Sayfa")) for r in sonuclar})
        kurtarma_yazi = f" ({kurtarilan} tanesi kurtarıldı)" if kurtarilan else ""

        gecmis_yazi = self.gecmise_kaydet(sonuclar)

        if durduruldu:
            self.bilgi.setText(
                f"⏹ Durduruldu — {sayfa_sayisi} sayfada {kimlik_sayisi} kimlik okundu"
                f"{kurtarma_yazi} (kalanlar işlenmedi).{gecmis_yazi}"
            )
        else:
            self.bilgi.setText(
                f"Tamamlandı — {sayfa_sayisi} sayfada {kimlik_sayisi} kimlik okundu"
                f"{kurtarma_yazi}.{gecmis_yazi}"
            )

        self.ozet_guncelle(sonuclar)
        if self.table.rowCount():
            self.table.selectRow(0)

    def gecmise_kaydet(self, sonuclar):
        """Tamamlanan taramayı kendiliğinden geçmişe yazar."""
        if not sonuclar:
            return ""
        try:
            self.gecmis_yolu = gecmis.taramayi_kaydet(
                sonuclar, kaynak_dosyalar=self.dosyalar
            )
        except Exception as exc:
            print(f"[GEÇMİŞ] kaydedilemedi: {exc!r}", flush=True)
            self.gecmis_yolu = None
            return "  (geçmişe kaydedilemedi)"
        return "  •  geçmişe kaydedildi" if self.gecmis_yolu else ""

    def gecmisi_guncellemeyi_planla(self):
        """Düzenleme sonrası geçmiş kaydını tazeler.

        Doğrudan yazmak yerine kısa bir gecikme kullanılıyor: hücre hücre
        düzenleme yapılırken her tuş için dosya yeniden yazılmasın, art arda
        gelen değişiklikler tek yazmada toplansın."""
        if self.gecmis_yolu and self.sonuclar:
            self.gecmis_zamanlayici.start()

    def gecmisi_guncelle(self):
        """Tablodaki güncel hali, açık olan geçmiş kaydının üzerine yazar."""
        self.gecmis_zamanlayici.stop()
        if not self.gecmis_yolu or not self.sonuclar:
            return
        try:
            # kaynak_dosyalar geçilmiyor: self.dosyalar kullanıcının BİR SONRAKİ
            # tarama için seçtiği dosyaları tutuyor olabilir; kaydın kendi dosya
            # listesi korunmalı.
            gecmis.taramayi_kaydet(self.sonuclar, yol=self.gecmis_yolu)
        except Exception as exc:
            print(f"[GEÇMİŞ] güncellenemedi: {exc!r}", flush=True)

    def listeyle_karsilastir(self):
        """Dışarıdan gelen listeyi yükleyip tablodaki sonuçlarla karşılaştırır."""
        if not self.sonuclar:
            return

        yol, _ = QFileDialog.getOpenFileName(
            self, "Karşılaştırılacak listeyi seç", "", "Excel (*.xlsx *.xlsm)"
        )
        if not yol:
            return

        try:
            kayitlar, bilgi = karsilastirma.listeyi_oku(yol)
        except Exception as exc:
            QMessageBox.warning(self, "Liste okunamadı", str(exc))
            return

        sonuc = karsilastirma.karsilastir(self.sonuclar, kayitlar)
        self.bilgi.setText(karsilastirma.ozet_metni(sonuc, os.path.basename(yol)))

        # Modal DEĞİL: kullanıcı bu pencere açıkken ana tabloda gezinebilsin.
        # Referansı sakla — yoksa show() sonrası nesne çöp toplanır ve pencere
        # hemen kapanır.
        self.karsilastirma_penceresi = KarsilastirmaPenceresi(
            sonuc, bilgi, os.path.basename(yol), self
        )
        self.karsilastirma_penceresi.show()
        self.karsilastirma_penceresi.raise_()
        self.karsilastirma_penceresi.activateWindow()

    def satiri_sec_ve_goster(self, satir):
        """Karşılaştırma penceresinden çağrılır: ana tabloda o satırı seçer,
        pencereyi öne getirir ve önizlemeyi gösterir."""
        try:
            r = next(i for i, s in enumerate(self.sonuclar) if s is satir)
        except StopIteration:
            return
        self.raise_()
        self.activateWindow()
        self.table.selectRow(r)
        self.table.scrollToItem(self.table.item(r, 0))

    def gecmisi_goster(self):
        pencere = GecmisPenceresi(self)
        if pencere.exec() == QDialog.Accepted and pencere.secilen_yol:
            self.gecmis_kaydini_ac(pencere.secilen_yol)

    def gecmis_kaydini_ac(self, yol):
        """Kayıtlı bir taramayı tabloya yükler."""
        if self.worker is not None and self.worker.isRunning():
            QMessageBox.information(
                self, "İşlem sürüyor",
                "Önce mevcut işlemin bitmesini bekleyin.",
            )
            return

        try:
            satirlar = gecmis.taramayi_yukle(yol)
        except Exception as exc:
            QMessageBox.critical(self, "Geçmiş açılamadı", f"{type(exc).__name__}: {exc}")
            return

        self.gecmisi_guncelle()      # açık kayıtta bekleyen değişiklik varsa önce onu yaz

        self.table.setRowCount(0)
        self.sonuclar = satirlar
        self.canli_sonuclar = list(satirlar)
        self.gecmis_yolu = yol
        self.pdf_bytes = None
        self._sayfa_onbellek = {}
        self._isaretli_onbellek = (None, None)

        self.tabloyu_yenile()
        self.ozet_guncelle(self.sonuclar)

        kimlik_sayisi = sum(1 for r in satirlar if r.get("_kart_bulundu"))
        sayfa_sayisi = len({(r.get("Dosya"), r.get("Sayfa")) for r in satirlar})
        self.bilgi.setText(
            f"Geçmişten açıldı — {sayfa_sayisi} sayfada {kimlik_sayisi} kimlik. "
            "Excel/PDF alabilir, satır ekleyip düzenleyebilirsiniz."
        )

        self.excel_btn.setEnabled(bool(satirlar))
        self.pdf_btn.setEnabled(bool(satirlar))
        self.edit_btn.setEnabled(bool(satirlar))
        self.karsilastir_btn.setEnabled(bool(satirlar))
        if self.table.rowCount():
            self.table.selectRow(0)

    def hata(self, mesaj):
        self.baslat_btn.setEnabled(True)
        QMessageBox.critical(self, "Hata", mesaj)

    def sayfa_goruntusu_getir(self, yol, sayfa_no):
        """Sayfanın tam görüntüsünü kaynak dosyadan üretir.

        Yüzlerce sayfanın görüntüsünü bellekte tutmak yerine, satır seçildiğinde
        yeniden okunuyor; son birkaç sayfa küçük bir önbellekte tutuluyor."""
        if not yol or not os.path.exists(yol):
            return None

        anahtar = (yol, sayfa_no)
        if anahtar in self._sayfa_onbellek:
            return self._sayfa_onbellek[anahtar]

        resim = None
        try:
            if excel_kaynak.excel_mi(yol):
                resim = excel_kaynak.bant_goruntusu(yol, sayfa_no)
            elif yol.lower().endswith(".pdf"):
                with pymupdf.open(yol) as doc:
                    indeks = (sayfa_no or 1) - 1
                    if 0 <= indeks < doc.page_count:
                        resim = pixmap_to_bgr(
                            doc[indeks].get_pixmap(dpi=PDF_RENDER_DPI, alpha=False)
                        )
            else:
                for _, img, _, _ in dosyayi_goruntulere_ayir(yol):
                    resim = img
                    break
        except Exception:
            return None

        if resim is not None:
            # Küçük bir FIFO önbellek: art arda aynı sayfanın satırları gezilirken
            # dosya tekrar tekrar açılmasın.
            if len(self._sayfa_onbellek) >= 3:
                self._sayfa_onbellek.pop(next(iter(self._sayfa_onbellek)))
            self._sayfa_onbellek[anahtar] = resim
        return resim

    def sayfanin_satirlari(self, sonuc):
        """Aynı dosya + sayfaya ait bütün satırlar (tablodaki sırayla)."""
        anahtar = (sonuc.get("Dosya"), sonuc.get("Sayfa"))
        return [r for r in self.sonuclar if (r.get("Dosya"), r.get("Sayfa")) == anahtar]

    def sayfayi_isaretle(self, sayfa_resmi, satirlar, secili):
        """Bulunan kimliklerin sayfa üzerindeki yerlerini numaralandırarak çizer."""
        img = sayfa_resmi.copy()
        h, w = img.shape[:2]
        kalinlik = max(2, int(min(h, w) * 0.004))
        yazi_olcek = max(0.8, min(h, w) / 900.0)

        for sira, satir in enumerate(satirlar, start=1):
            koseler = satir.get("_koseler")
            if not koseler:
                continue

            secili_mi = satir is secili
            if satir.get("_manuel_eklendi"):
                renk = (255, 170, 60)
            elif satir.get("_kurtarildi"):
                renk = (0, 170, 255)
            else:
                renk = (80, 220, 80)

            poly = np.asarray(koseler, dtype=np.int32).reshape(-1, 1, 2)
            cv2.polylines(img, [poly], True, renk,
                          kalinlik * (2 if secili_mi else 1), cv2.LINE_AA)

            x, y = poly[0][0]
            etiket = str(sira)
            (tw, th), _ = cv2.getTextSize(etiket, cv2.FONT_HERSHEY_SIMPLEX, yazi_olcek, 2)
            x = int(max(0, min(x, w - tw - 12)))
            y = int(max(th + 12, min(y, h - 6)))
            cv2.rectangle(img, (x, y - th - 10), (x + tw + 12, y + 4), renk, -1)
            cv2.putText(img, etiket, (x + 6, y - 4), cv2.FONT_HERSHEY_SIMPLEX,
                        yazi_olcek, (20, 20, 20), 2, cv2.LINE_AA)

        return img

    def onizleme_gorselleri(self, sonuc):
        """(gösterilecek görüntü, işaretsiz görüntü, sayfadaki ofset) üçlüsü.

        "kart" kipinde (bir satıra tıklandığında) kartın PERSPEKTİFİ
        DÜZELTİLMİŞ, düz hali (_preview) gösterilir — sayfadan kırpma
        yapılmaz, komşu kartlar veya kâğıdın eğikliği görünmez. "sayfa"
        kipinde (Sayfa sütununa tıklandığında) sayfanın tamamı, bulunan
        kartlar işaretli halde gösterilir; fareyle seçilen alan bu işaretsiz
        sayfadan kırpılır (yeşil çerçeveler OCR'a karışmasın)."""
        if self.coklu_cb.isChecked() and self._onizleme_kipi == "kart":
            duz = sonuc.get("_preview")
            if duz is not None:
                return duz, duz, (0, 0)

            # Kart hizalanamadı (kurtarma/elle açılmış boş satır gibi):
            # elimizde düz bir görüntü yok, sayfadan kimliğin çevresini
            # kırpıp gösteriyoruz — bu durumda seçim yine sayfa koordinatında.
            sayfa = self.sayfa_goruntusu_getir(sonuc.get("_kaynak_yol"), sonuc.get("Sayfa"))
            if sayfa is not None and sonuc.get("_koseler"):
                kirpma, ofset = self.karta_yakinlas(sayfa, sonuc["_koseler"])
                return kirpma, kirpma, ofset
            if sayfa is not None:
                return sayfa, sayfa, (0, 0)

        elif self.coklu_cb.isChecked():   # "sayfa" kipi: tüm sayfa, işaretli
            sayfa = self.sayfa_goruntusu_getir(
                sonuc.get("_kaynak_yol"), sonuc.get("Sayfa")
            )
            if sayfa is not None:
                satirlar = self.sayfanin_satirlari(sonuc)
                anahtar = (sonuc.get("_kaynak_yol"), sonuc.get("Sayfa"), id(sonuc),
                           tuple(id(r) for r in satirlar))
                if self._isaretli_onbellek[0] == anahtar:
                    return self._isaretli_onbellek[1]

                isaretli = self.sayfayi_isaretle(sayfa, satirlar, sonuc)
                gorseller = (isaretli, sayfa, (0, 0))
                self._isaretli_onbellek = (anahtar, gorseller)
                return gorseller

        if self.debug_cb.isChecked():
            for anahtar in ("_debug_resmi", "_kart_debug"):
                if sonuc.get(anahtar) is not None:
                    return sonuc[anahtar], sonuc[anahtar], (0, 0)
        onizleme = sonuc.get("_preview")
        return onizleme, onizleme, (0, 0)

    def etiketi_sigdir(self, etiket, tam_metin):
        """Tek satırlık etikete sığmayan metni "…" ile kısaltarak yazar."""
        genislik = max(60, etiket.width() - 4)
        etiket.setText(etiket.fontMetrics().elidedText(tam_metin, Qt.ElideRight, genislik))
        etiket.setToolTip(tam_metin if tam_metin else "")

    def onizleme_yazilarini_tazele(self):
        self.etiketi_sigdir(self.sayfa_bilgi, self._sayfa_bilgi_metni)
        self.etiketi_sigdir(self.onizleme_ipucu, self._onizleme_ipucu_metni)

    def sayfa_bilgisini_guncelle(self, sonuc):
        if not self.coklu_cb.isChecked():
            self._sayfa_bilgi_metni = ""
            self._onizleme_ipucu_metni = ""
            self.onizleme_yazilarini_tazele()
            return

        satirlar = self.sayfanin_satirlari(sonuc)
        bulunan = sum(1 for r in satirlar if r.get("_kart_bulundu"))
        elle = sum(1 for r in satirlar if r.get("_manuel_eklendi"))
        kurtarilan = sum(1 for r in satirlar if r.get("_kurtarildi"))

        sayfa_yazi = f"Sayfa {sonuc.get('Sayfa')}" if sonuc.get("Sayfa") != "-" else "Görüntü"
        metin = f"{sayfa_yazi} — {bulunan} kimlik" if bulunan else f"{sayfa_yazi} — kimlik yok"

        notlar = []
        if kurtarilan:
            notlar.append(f"{kurtarilan} kurtarıldı")
        if elle:
            notlar.append(f"{elle} elle eklendi")
        if notlar:
            metin += " (" + ", ".join(notlar) + ")"

        kart_kipi = self._onizleme_kipi == "kart"
        if kart_kipi and sonuc.get("_preview") is not None:
            metin += "  •  düzeltilmiş kart"
        elif kart_kipi and sonuc.get("_koseler"):
            metin += "  •  yakınlaştırıldı"

        self._sayfa_bilgi_metni = metin
        if kart_kipi and sonuc.get("_preview") is not None:
            # Bu görünümde alan seçimi kapalı (koordinat sayfaya değil kartın
            # kendi pikseline ait olurdu); ipucu buna göre kısaltılıyor.
            self._onizleme_ipucu_metni = "Sayfanın tamamı için Sayfa sütununa tıklayın."
        else:
            self._onizleme_ipucu_metni = (
                "Sayfanın tamamı için Sayfa sütununa tıklayın.  "
                "İşaretlenmemiş kimliği fareyle çerçeveleyip “Seçili alanı oku” ile okutun."
            )
        self.onizleme_yazilarini_tazele()

    def pixmapi_yerlestir(self, img, ham=None, ofset=(0, 0), secim_acik=True):
        self.preview.secimi_temizle()
        self.preview.secim_acik = secim_acik
        self._secim = None
        self.alan_oku_btn.setEnabled(False)
        self._onizleme_ham = ham if ham is not None else img
        self._onizleme_ofset = ofset

        if img is None:
            self.preview.clear()
            self.preview.setMinimumSize(320, 220)
            self.preview.setText("Görüntü bulunamadı.")
            return

        if self.zoom is None:
            alan_w = max(260, self.preview_scroll.viewport().width() - 8)
            alan_h = max(180, self.preview_scroll.viewport().height() - 8)
            self.preview.setMinimumSize(1, 1)
            pixmap = bgr_to_pixmap(img, max_w=alan_w, max_h=alan_h)
        else:
            h, w = img.shape[:2]
            # Ölçeklenmiş görüntü belleği patlatmasın: kenar en fazla 6000 piksel.
            azami_olcek = 6000.0 / max(h, w)
            self.zoom = min(self.zoom, azami_olcek)
            pixmap = bgr_to_pixmap(img, max_w=int(w * self.zoom), max_h=int(h * self.zoom))
            self.preview.setMinimumSize(pixmap.size())

        self.preview.setPixmap(pixmap)
        self.preview.setAlignment(Qt.AlignCenter)
        # Pixmap ile görüntü arasındaki oran: seçimi görüntü koordinatına çevirmek için.
        self._onizleme_olcek = pixmap.width() / float(img.shape[1]) if img.shape[1] else 1.0

    def alan_secildi(self, kutu):
        """Önizlemede fareyle çizilen dikdörtgeni görüntü koordinatına çevirir."""
        if kutu is None or self._onizleme_ham is None or self._onizleme_olcek <= 0:
            self._secim = None
            self.alan_oku_btn.setEnabled(False)
            return

        h, w = self._onizleme_ham.shape[:2]
        olcek = self._onizleme_olcek
        x0 = max(0, min(w - 1, int(kutu.left() / olcek)))
        y0 = max(0, min(h - 1, int(kutu.top() / olcek)))
        x1 = max(0, min(w, int(kutu.right() / olcek)))
        y1 = max(0, min(h, int(kutu.bottom() / olcek)))

        if x1 - x0 < 40 or y1 - y0 < 25:
            self._secim = None
            self.alan_oku_btn.setEnabled(False)
            self._sayfa_bilgi_metni = "Seçilen alan çok küçük — kimliği tamamen çerçeveleyin."
            self.onizleme_yazilarini_tazele()
            return

        self._secim = (x0, y0, x1, y1)
        self.alan_oku_btn.setEnabled(self.alan_okuyucu is None)

    def secili_alani_oku(self):
        """Kullanıcının çerçevelediği alanı okuyup yeni bir satır olarak ekler."""
        if self._secim is None or self._onizleme_ham is None:
            return
        if self.alan_okuyucu is not None:
            return

        r = self.table.currentRow()
        if r < 0 or r >= len(self.sonuclar):
            return

        x0, y0, x1, y1 = self._secim
        kirpma = np.ascontiguousarray(self._onizleme_ham[y0:y1, x0:x1])
        if kirpma.size == 0:
            return

        self._okunan_satir = r
        self._okunan_kutu = self._secim
        self._okunan_ofset = self._onizleme_ofset
        self.alan_oku_btn.setEnabled(False)
        self.alan_oku_btn.setText("Okunuyor…")
        self._sayfa_bilgi_metni = "Seçilen alan okunuyor…"
        self.onizleme_yazilarini_tazele()

        self.alan_okuyucu = AlanOkuyucu(kirpma, debug=self.debug_cb.isChecked())
        self.alan_okuyucu.bitti.connect(self.alan_okundu)
        self.alan_okuyucu.hata.connect(self.alan_okuma_hatasi)
        self.alan_okuyucu.finished.connect(self.alan_okuma_bitti)
        self.alan_okuyucu.start()

    def alan_okuma_bitti(self):
        self.alan_okuyucu = None
        self.alan_oku_btn.setText("🔍 Seçili alanı oku")
        self.alan_oku_btn.setEnabled(self._secim is not None)

    def alan_okuma_hatasi(self, mesaj):
        QMessageBox.warning(self, "Alan okunamadı", mesaj)

    def alan_okundu(self, sonuc):
        """Seçili alandan okunanları yeni bir satır olarak tabloya ekler."""
        kaynak = self.sonuclar[self._okunan_satir]
        ocr_sonuc = sonuc.get("ocr", {}) or {}
        # Seçim, ekranda gösterilen (belki kırpılmış) görüntünün koordinatındaydı;
        # işaretlemenin sayfada doğru yere düşmesi için ofset geri ekleniyor.
        ox, oy = self._okunan_ofset
        x0, y0, x1, y1 = self._okunan_kutu
        x0, y0, x1, y1 = x0 + ox, y0 + oy, x1 + ox, y1 + oy

        kimlik_no = str(ocr_sonuc.get("tc_no", "Bulunamadi")).strip() or "Bulunamadi"
        ad = ocr_sonuc.get("ad", "Bulunamadi")
        soyad = ocr_sonuc.get("soyad", "Bulunamadi")
        tc_supheli = kimlik_no != "Bulunamadi" and not ocr_sonuc.get("tc_dogrulandi", True)

        eksikler = [
            alan for alan, deger in
            (("Kimlik No", kimlik_no), ("Ad", ad), ("Soyad", soyad))
            if deger == "Bulunamadi"
        ]
        if eksikler:
            durum = "Seçilen alandan okundu — eksik: " + ", ".join(eksikler)
        elif tc_supheli:
            durum = "Seçilen alandan okundu — Kimlik No doğrulanamadı"
        else:
            durum = "Seçilen alandan okundu"

        belge_yazi = {
            "tc": "T.C. Kimlik", "eski_tc": "Eski T.C. Kimlik", "gocmen": "Göçmen / Yabancı",
        }.get(sonuc.get("belge_tipi", "tc"), "T.C. Kimlik")

        yeni = {
            "Dosya": kaynak.get("Dosya", ""),
            "Sayfa": kaynak.get("Sayfa", "-"),
            "Kart": "seçim",
            "Belge Türü": belge_yazi,
            "Kimlik No": kimlik_no,
            "Ad": ad,
            "Soyad": soyad,
            "Bitiş Tarihi": ocr_sonuc.get("bitis_tarihi", "") or "-",
            "Geçerlilik": "-",
            "Durum": durum,
            "Süre": round(float(ocr_sonuc.get("ocr_suresi", 0.0) or 0.0), 2),
            "_belge_tipi": sonuc.get("belge_tipi", "tc"),
            "_belge_gecerli": ocr_sonuc.get("belge_gecerli"),
            "_kart_bulundu": True,
            "_manuel_eklendi": True,
            "_tc_supheli": tc_supheli,
            "_kaynak_yol": kaynak.get("_kaynak_yol"),
            "_koseler": [[x0, y0], [x1, y0], [x1, y1], [x0, y1]],
            "_preview": sonuc.get("kart"),
            "_debug_resmi": ocr_sonuc.get("debug_resmi"),
            "_kart_sonuc": {
                "belge_tipi": sonuc.get("belge_tipi"),
                "arama_modu": "elle seçim",
                "mesaj": ("Seçilen alanda kart hizalandı."
                          if sonuc.get("hizalandi") else "Seçilen alan doğrudan okundu."),
                "referans_hatalari": {},
            },
            "_ocr_sonuc": {
                "guven": ocr_sonuc.get("guven"),
                "tc_dogrulandi": ocr_sonuc.get("tc_dogrulandi"),
                "tc_kaynak": ocr_sonuc.get("tc_kaynak"),
                "kimlik_no_conf": ocr_sonuc.get("kimlik_no_conf", 0.0),
                "ad_conf": ocr_sonuc.get("ad_conf", 0.0),
                "soyad_conf": ocr_sonuc.get("soyad_conf", 0.0),
                "ocr_suresi": ocr_sonuc.get("ocr_suresi", 0.0),
                "tum_ocr": ocr_sonuc.get("tum_ocr", []),
            },
        }

        # Aynı sayfanın son satırından sonraya ekle; sıralama ve numaralandırma
        # zaten konuma göre yeniden yapılacak.
        anahtar = (yeni["Dosya"], yeni["Sayfa"])
        hedef = len(self.sonuclar)
        for i in range(len(self.sonuclar) - 1, -1, -1):
            if (self.sonuclar[i].get("Dosya"), self.sonuclar[i].get("Sayfa")) == anahtar:
                hedef = i + 1
                break

        self.sonuclar.insert(hedef, yeni)
        self.sayfalari_duzenle()
        self.tabloyu_yenile(secilecek=yeni)
        self.ozet_guncelle(self.sonuclar)
        self.gecmisi_guncellemeyi_planla()

    def zoom_degistir(self, carpan):
        if self.zoom is None:
            # Sığdırma modundayken mevcut görünen boyuttan devam et.
            pixmap = self.preview.pixmap()
            img = self.aktif_onizleme_resmi()
            if pixmap is None or img is None:
                return
            self.zoom = pixmap.width() / float(img.shape[1])
        self.zoom = max(0.1, min(8.0, self.zoom * carpan))
        self.onizleme_goster()

    def zoom_sigdir(self):
        self.zoom = None
        self.onizleme_goster()

    def aktif_onizleme_resmi(self):
        r = self.table.currentRow()
        if r < 0 or r >= len(self.sonuclar):
            return None
        return self.onizleme_gorselleri(self.sonuclar[r])[0]

    def eventFilter(self, nesne, olay):
        """Ctrl + fare tekerleği ile yakınlaştırma."""
        if (nesne is self.preview_scroll.viewport()
                and olay.type() == QEvent.Wheel
                and olay.modifiers() & Qt.ControlModifier):
            self.zoom_degistir(1.15 if olay.angleDelta().y() > 0 else 1 / 1.15)
            return True
        return super().eventFilter(nesne, olay)

    def sayfa_satirlarini_sirala(self, satirlar):
        """Bir sayfanın satırlarını kimliğin sayfadaki YERİNE göre dizer:
        yukarıdan aşağıya, aynı hizadakiler soldan sağa. Konumu olmayan
        (elle açılmış boş) satırlar sona alınır."""
        konumlu = [r for r in satirlar if r.get("_koseler")]
        konumsuz = [r for r in satirlar if not r.get("_koseler")]

        if len(konumlu) > 1:
            # Tespit tarafındaki sıralama ile aynı kural kullanılsın diye
            # satırlar geçici olarak o fonksiyonun beklediği şekle sarılıyor.
            sarmal = [{"koseler": np.asarray(r["_koseler"], dtype=np.float32), "_satir": r}
                      for r in konumlu]
            konumlu = [k["_satir"] for k in sayfa_sirasina_diz(sarmal)]

        if not konumsuz:
            return konumlu

        # Konumu olmayan (elle açılmış boş) satırlar, açıldıkları satırın hemen
        # ardına yerleşir; böylece numarası da oradan devam eder. Dayandığı
        # satır kalmadıysa sayfanın sonuna eklenir.
        sonuc = list(konumlu)
        artakalan = []
        for satir in konumsuz:
            ardil = satir.get("_ardil_oldugu")
            yerlesti = False
            if ardil is not None:
                for i, mevcut in enumerate(sonuc):
                    if id(mevcut) == ardil:
                        sonuc.insert(i + 1, satir)
                        yerlesti = True
                        break
            if not yerlesti:
                artakalan.append(satir)

        return sonuc + artakalan

    def sayfalari_duzenle(self):
        """Tüm satırları dosya+sayfa gruplarında konuma göre sıralar ve Kart
        numaralarını (1/4, 2/4 …) baştan verir.

        Araya elle bir kimlik eklendiğinde numarasını konumundan alır,
        kendisinden sonraki kimliklerin numarası da bir artar."""
        gruplar = {}
        sira = []
        for satir in self.sonuclar:
            anahtar = (satir.get("Dosya"), satir.get("Sayfa"))
            if anahtar not in gruplar:
                gruplar[anahtar] = []
                sira.append(anahtar)
            gruplar[anahtar].append(satir)

        yeni_liste = []
        for anahtar in sira:
            sirali = self.sayfa_satirlarini_sirala(gruplar[anahtar])
            toplam = sum(1 for r in sirali if r.get("_kart_bulundu"))
            sayac = 0
            for satir in sirali:
                if satir.get("_kart_bulundu"):
                    sayac += 1
                    satir["Kart"] = f"{sayac}/{toplam}"
                else:
                    satir["Kart"] = "-"
            yeni_liste.extend(sirali)

        self.sonuclar = yeni_liste
        self.canli_sonuclar = list(self.sonuclar)
        self._isaretli_onbellek = (None, None)

    def tabloyu_yenile(self, secilecek=None):
        """Tabloyu self.sonuclar'dan baştan kurar (sıra değiştiğinde).
        `secilecek` verilen satır sözlüğü tekrar seçili hale gelir."""
        self._table_updating = True
        try:
            self.table.setRowCount(0)
            for satir in self.sonuclar:
                self.tabloya_satir_yaz(satir)
        finally:
            self._table_updating = False

        if secilecek is not None:
            for i, satir in enumerate(self.sonuclar):
                if satir is secilecek:
                    self.table.selectRow(i)
                    break

    @staticmethod
    def duzenlenmis_hucreyi_boya(item):
        """Elle değiştirilen hücrenin vurgusu.

        Eskiden açık mavi bir ZEMİN veriliyordu; koyu temada yazı da açık
        olduğu için hücre okunmaz hale geliyordu. Artık koyu lacivert zemin +
        açık mavi yazı kullanılıyor."""
        item.setBackground(QColor("#1f3350"))
        item.setForeground(QColor("#93c5fd"))

    def tabloya_satir_yaz(self, row, hedef=None):
        """Tek bir satırı tabloya ekler (renklendirme kurallarıyla birlikte)."""
        r = self.table.rowCount() if hedef is None else hedef
        self.table.insertRow(r)

        editable_headers = {"Kimlik No", "Ad", "Soyad"}
        for c, h in enumerate(self.headers):
            item = QTableWidgetItem(str(row.get(h, "")))

            if self.edit_mode and h in editable_headers:
                item.setFlags(item.flags() | Qt.ItemIsEditable)
            else:
                item.setFlags(item.flags() & ~Qt.ItemIsEditable)

            if row.get("_belge_gecerli") is False and h == "Geçerlilik":
                item.setForeground(QColor("#ff6b6b"))
            if row.get("_kurtarildi") and h in ("Kart", "Durum"):
                item.setForeground(QColor("#f0b429"))
            if row.get("_tc_supheli") and h in ("Kimlik No", "Durum"):
                item.setForeground(QColor("#f0b429"))
            if row.get("_manuel_eklendi"):
                item.setForeground(QColor("#ffb454"))
            # Excel'den tamamlanan ad/soyad ve ad uyuşmazlığı
            if h in (row.get("_excelden_alanlar") or ()):
                item.setForeground(QColor("#a5b4fc"))
            if row.get("_ad_uyusmazligi") and h in ("Ad", "Soyad", "Durum"):
                item.setForeground(QColor("#f0b429"))
            if h in (row.get("_duzenlenen_alanlar") or ()):
                self.duzenlenmis_hucreyi_boya(item)

            self.table.setItem(r, c, item)
        return r

    def hucreye_tiklandi(self, satir, sutun):
        """Sayfa sütunu tüm sayfayı, diğer sütunlar o satırın kimliğini gösterir."""
        yeni_kip = "sayfa" if self.headers[sutun] == "Sayfa" else "kart"
        if yeni_kip != self._onizleme_kipi:
            self._onizleme_kipi = yeni_kip
            self.zoom = None          # kip değişince sığdırmaya dön
        self.onizleme_goster()

    def karta_yakinlas(self, resim, koseler, pay_orani=0.18):
        """Sayfa görüntüsünü, ilgili kimliğin çevresinden kırpar.
        Dönüş: (kırpılmış görüntü, (ofset_x, ofset_y))"""
        h, w = resim.shape[:2]
        nokta = np.asarray(koseler, dtype=np.float32).reshape(-1, 2)
        x0, y0 = nokta[:, 0].min(), nokta[:, 1].min()
        x1, y1 = nokta[:, 0].max(), nokta[:, 1].max()

        pay = pay_orani * max(x1 - x0, y1 - y0)
        x0 = int(max(0, x0 - pay))
        y0 = int(max(0, y0 - pay))
        x1 = int(min(w, x1 + pay))
        y1 = int(min(h, y1 + pay))

        if x1 - x0 < 20 or y1 - y0 < 20:
            return resim, (0, 0)
        return np.ascontiguousarray(resim[y0:y1, x0:x1]), (x0, y0)

    def onizleme_goster(self):
        r = self.table.currentRow()
        if r < 0 or r >= len(self.sonuclar):
            return

        sonuc = self.sonuclar[r]

        gosterilecek, ham, ofset = self.onizleme_gorselleri(sonuc)
        # "kart" kipinde perspektifi düzeltilmiş kartın kendisi gösteriliyor;
        # üzerinde fareyle alan seçmenin bir anlamı yok (koordinat sayfaya
        # değil kartın kendi pikseline ait olurdu) — o yüzden kapatılıyor.
        secim_acik = not (
            self.coklu_cb.isChecked()
            and self._onizleme_kipi == "kart"
            and sonuc.get("_preview") is not None
        )
        self.pixmapi_yerlestir(gosterilecek, ham=ham, ofset=ofset, secim_acik=secim_acik)
        self.sayfa_bilgisini_guncelle(sonuc)

        if not self.debug_cb.isChecked():
            self.debug_text.clear()
            return

        kart = sonuc.get("_kart_sonuc", {}) or {}
        ocr = sonuc.get("_ocr_sonuc", {}) or {}

        satirlar = [
            f"Dosya: {sonuc.get('Dosya')}",
            f"Sayfa: {sonuc.get('Sayfa')}",
            f"Kart: {sonuc.get('Kart')}",
            f"Belge Türü: {sonuc.get('Belge Türü')}",
            "",
            "=== SONUÇ ===",
            f"Kimlik No: {sonuc.get('Kimlik No')}",
            f"Ad: {sonuc.get('Ad')}",
            f"Soyad: {sonuc.get('Soyad')}",
            f"Bitiş Tarihi: {sonuc.get('Bitiş Tarihi')}",
            f"Geçerlilik: {sonuc.get('Geçerlilik')}",
            f"Durum: {sonuc.get('Durum')}",
            "",
            "=== BELGE TESPİT ===",
            f"Seçilen tip: {kart.get('belge_tipi')}",
            f"Düzeltildi: {kart.get('duzeltildi')}",
            f"Düzeltme yöntemi: {kart.get('duzeltme_yontemi')}",
            f"Fallback: {kart.get('fallback')}",
            f"İyi eşleşme: {kart.get('iyi_eslesme')}",
            f"Inlier: {kart.get('inlier')}",
            f"Inlier oranı: {kart.get('inlier_orani')}",
            f"Skor: {kart.get('skor')}",
            f"Sayfadaki kart: {kart.get('kart_sirasi')}/{kart.get('kart_sayisi')}",
            f"Arama modu: {kart.get('arama_modu')}",
            f"Mesaj: {kart.get('mesaj')}",
            f"Aday sırası: {kart.get('aday_sirasi')}",
            "",
            "=== OCR ===",
            f"Güven: {ocr.get('guven')}",
            f"Kimlik No doğrulandı: {ocr.get('tc_dogrulandi')} (kaynak: {ocr.get('tc_kaynak')})",
            f"Kimlik no confidence: {float(ocr.get('kimlik_no_conf', 0.0) or 0.0):.3f}",
            f"Ad confidence: {float(ocr.get('ad_conf', 0.0) or 0.0):.3f}",
            f"Soyad confidence: {float(ocr.get('soyad_conf', 0.0) or 0.0):.3f}",
            f"OCR süresi: {float(ocr.get('ocr_suresi', 0.0) or 0.0):.3f} sn",
        ]

        if sonuc.get("_belge_tipi") == "gocmen":
            satirlar.extend([
                "",
                "=== GÖÇMEN OCR ===",
                f"Başlangıç: {ocr.get('baslangic_tarihi')}",
                f"Bitiş: {ocr.get('bitis_tarihi')}",
                f"Geçerlilik durumu: {ocr.get('gecerlilik_durumu')}",
                f"Hızlı yol: {ocr.get('hizli_yol_kullanildi')}",
                f"Hızlı hücre OCR: {float(ocr.get('hizli_hucre_ocr_suresi', 0.0) or 0.0):.3f} sn",
                f"Detector fallback: {ocr.get('detector_fallback_kullanildi')}",
                f"Detector OCR: {float(ocr.get('detector_ocr_suresi', 0.0) or 0.0):.3f} sn",
                f"Fallback denenen: {ocr.get('fallback_denenen_alanlar')}",
                f"Fallback kullanılan: {ocr.get('fallback_kullanilan_alanlar')}",
                f"Fallback OCR süresi: {float(ocr.get('fallback_ocr_suresi', 0.0) or 0.0):.3f} sn",
            ])

        ham = ocr.get("tum_ocr", []) or []
        if ham:
            satirlar.append("")
            satirlar.append("=== HAM OCR ===")
            for item in ham[:80]:
                try:
                    satirlar.append(
                        f"{item.get('text', '')} — {float(item.get('conf', 0.0) or 0.0):.3f}"
                    )
                except Exception:
                    satirlar.append(str(item))

        self.debug_text.setPlainText(
            "\n".join(satirlar)
        )


    def excel_kaydet(self):
        """
        Sonuçları geçici bir Excel dosyasına çevirir ve
        varsayılan Excel uygulamasıyla doğrudan açar.

        Kullanıcı isterse Excel içinden normal "Farklı Kaydet"
        ile kalıcı konum seçebilir.
        """
        if not self.sonuclar:
            return

        try:
            fd, yol = tempfile.mkstemp(
                prefix="kimlik_sonuclari_",
                suffix=".xlsx",
            )
            os.close(fd)

            # Nihai Excel sadece:
            # Kimlik No | Ad | Soyad
            df = pd.DataFrame([
                {
                    "Kimlik No": r["Kimlik No"],
                    "Ad": r["Ad"],
                    "Soyad": r["Soyad"],
                }
                for r in self.sonuclar
            ])

            df.to_excel(
                yol,
                index=False,
                engine="openpyxl",
            )

            wb = load_workbook(
                yol
            )

            ws = wb.active
            ws.title = "Kimlik Sonuçları"

            # Excel'de GEÇERSİZ kimlikler ve checksum'ı tutmayan
            # (doğrulanamamış) kimlik numaraları renklendirilir; ikincisi
            # kullanıcının gözden geçirmesi gereken tek şey.
            gecersiz_fill = PatternFill(
                "solid",
                fgColor="FFF2CC",
            )

            supheli_fill = PatternFill(
                "solid",
                fgColor="FFD9A0",
            )

            for cell in ws[1]:
                cell.font = Font(
                    bold=True
                )

            for i, r in enumerate(
                self.sonuclar,
                start=2,
            ):
                if r.get(
                    "_belge_gecerli"
                ) is False:
                    # Excel yalnız 3 kolon: Kimlik No | Ad | Soyad
                    # Geçersiz belgenin tamamı tek renkle görünür.
                    for col in range(
                        1,
                        4,
                    ):
                        ws.cell(
                            i,
                            col,
                        ).fill = gecersiz_fill

                if r.get("_tc_supheli"):
                    hucre = ws.cell(i, 1)
                    hucre.fill = supheli_fill
                    hucre.comment = Comment(
                        "Bu numara okundu ama doğrulama hanesi tutmuyor; "
                        "kimlikle karşılaştırıp kontrol edin.",
                        "Kimlik Okuyucu",
                    )

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 24
            ws.column_dimensions["C"].width = 24

            wb.save(
                yol
            )

            dosyayi_sistemde_ac(
                yol
            )

            self.bilgi.setText(
                "Excel oluşturuldu ve açıldı. "
                "Kalıcı olarak saklamak istersen Excel içinden Farklı Kaydet kullanabilirsin."
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Excel oluşturma hatası",
                str(e),
            )


    def pdf_olustur(self):
        """Tablodaki GÜNCEL sıraya göre kimlik PDF'i üretir.

        Eskiden PDF işlem sırasında hazırlanıyordu; elle kesilen/eklenen
        kimlikler ve sıralama değişiklikleri içine girmiyordu. Artık kaydetme
        anında tablodan üretiliyor: her satırın kendi görüntüsü, tabloda
        göründüğü sırayla (sayfa sayfa, her sayfada yukarıdan aşağıya) eklenir.
        """
        doc = pymupdf.open()
        try:
            for satir in self.sonuclar:
                resim = satir.get("_preview")
                if resim is not None:
                    resmi_pdfe_ekle(doc, resim)
            if not doc.page_count:
                return None
            return doc.tobytes(garbage=3, deflate=True)
        finally:
            doc.close()

    def pdf_kaydet(self):
        """
        Kimlik PDF'ini geçici dosyaya yazar ve varsayılan PDF
        görüntüleyiciyle doğrudan açar.
        """
        if not self.sonuclar:
            return

        self.pdf_bytes = self.pdf_olustur()
        if not self.pdf_bytes:
            QMessageBox.information(
                self,
                "PDF oluşturulamadı",
                "Tabloda PDF'e eklenecek görüntü bulunamadı.",
            )
            return

        try:
            fd, yol = tempfile.mkstemp(
                prefix="kimlikler_",
                suffix=".pdf",
            )
            os.close(fd)

            with open(
                yol,
                "wb",
            ) as f:
                f.write(
                    self.pdf_bytes
                )

            dosyayi_sistemde_ac(
                yol
            )

            self.bilgi.setText(
                "PDF oluşturuldu ve açıldı. "
                "Kalıcı olarak saklamak istersen PDF görüntüleyicinden kaydedebilirsin."
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "PDF oluşturma hatası",
                str(e),
            )


    def closeEvent(self, event):
        """Pencere kapanırken süren tarama arka planda devam etmesin.
        Worker o an işlediği sayfayı bitirip döngüden çıkar."""
        if self.worker is not None and self.worker.isRunning():
            self.worker.durdur()
            self.worker.wait(5000)

        # Zamanlayıcıda bekleyen düzenleme varsa diske yaz.
        if self.gecmis_zamanlayici.isActive():
            self.gecmisi_guncelle()

        super().closeEvent(event)

    def resizeEvent(self, event):
        super().resizeEvent(event)

        if hasattr(self, "sayfa_bilgi"):
            self.onizleme_yazilarini_tazele()

        # Pencere boyutu değişince seçili kimliği yeni alana tekrar sığdır.
        if hasattr(self, "table") and self.table.currentRow() >= 0:
            self.onizleme_goster()



if __name__ == "__main__":
    print("[CHECK] desktop.py içinde cv2.imdecode kullanılmıyor.")
    app = QApplication(sys.argv)
    pencere = MainWindow()
    pencere.showMaximized()
    sys.exit(app.exec())
